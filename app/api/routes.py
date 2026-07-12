from datetime import date, datetime, timedelta
from hashlib import sha256
from io import BytesIO
import json
import secrets
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import and_, func, or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.config import settings
from app.core.rbac import (
    Actor,
    get_current_actor,
    require_bound_device,
    require_platform_admin,
    require_roles,
    require_work_order_execution_scope,
    require_work_order_owner_scope,
    require_work_order_scope,
    require_work_order_write_scope,
)
from app.core.security import create_access_token, hash_password, verify_password
from app.models import (
    AuditLog,
    CompletionPolicy,
    Customer,
    Equipment,
    InventoryTransaction,
    InventoryNotification,
    ReplenishmentRequest,
    VehicleReturnRequest,
    ImportBatch,
    JobStatus,
    Organization,
    Part,
    PartMachineAssociation,
    WorkOrderPartMemory,
    QCPicture,
    ReturnEquipment,
    StorageLocation,
    TransactionType,
    User,
    UserDevice,
    UserInvitation,
    UserRole,
    Warehouse,
    WorkOrder,
    WorkOrderPart,
    WorkOrderVoiceNote,
)
from app.schemas import (
    AbnormalUsageRow,
    CustomerCreate,
    CustomerRead,
    EquipmentCreate,
    EquipmentRead,
    CompletionPolicyRead,
    CompletionPolicyUpsert,
    JobStatusCreate,
    JobStatusRead,
    QCPictureCreate,
    QCPictureRead,
    ReturnEquipmentCreate,
    ReturnEquipmentRead,
    InventoryTransactionCreate,
    LowStockAlert,
    LocationStockBalance,
    InventoryScanRequest,
    InventoryScanRead,
    WorkOrderFlowAction,
    InventoryTransactionRead,
    ImportBatchRead,
    InvitationAccept,
    InvitationCreate,
    InvitationCreated,
    InvitationInfo,
    AdminWarehouseDashboard,
    EngineerDashboard,
    PartCreate,
    PartRead,
    PartMachineAssociationRead,
    WorkOrderPartRecommendation,
    InventoryNotificationRead,
    ReplenishmentRequestRead,
    ReplenishmentRequestAction,
    ReplenishmentRequestCreate,
    ReplenishmentRequestReconcile,
    VehicleReturnRequestAction,
    VehicleReturnRequestCreate,
    VehicleReturnRequestRead,
    OrganizationCreate,
    OrganizationRead,
    OrganizationUpdate,
    PasswordSet,
    StockBalance,
    StorageLocationCreate,
    StorageLocationRead,
    TokenResponse,
    UserCreate,
    UserRead,
    WarehouseCreate,
    WarehouseRead,
    WorkOrderCreate,
    WorkOrderClaimRelease,
    WorkOrderPartCreate,
    WorkOrderPartRead,
    WorkOrderProfit,
    WorkOrderRead,
    WorkOrderUpdate,
    WorkOrderServiceContext,
    WorkOrderVoiceNoteRead,
    WarehouseSummary,
)
from app.services.inventory import (
    create_transaction,
    get_employee_van_inventory,
    get_stock_quantity,
    get_available_stock_quantity,
    get_stock_balances,
    get_location_stock_balances,
    get_location_stock_quantity,
    get_work_order_parts_cost,
    use_part_on_work_order,
    warehouse_is_vehicle,
    begin_inventory_write,
)

router = APIRouter()

PART_IMPORT_FIELDS = {
    "part_number",
    "name",
    "category",
    "barcode",
    "item_type",
    "tracking_mode",
    "is_active",
    "english_name",
    "machine_type",
    "unit",
    "default_cost",
    "safety_stock",
    "min_stock",
    "supplier",
    "image_url",
    "notes",
}


def _import_batch_read(batch: ImportBatch) -> ImportBatchRead:
    payload = json.loads(batch.payload_json or "[]")
    return ImportBatchRead(
        id=batch.id,
        organization_id=batch.organization_id,
        import_type=batch.import_type,
        filename=batch.filename,
        file_sha256=batch.file_sha256,
        status=batch.status,
        total_rows=batch.total_rows,
        valid_rows=batch.valid_rows,
        error_rows=batch.error_rows,
        created_count=batch.created_count,
        updated_count=batch.updated_count,
        errors=json.loads(batch.errors_json or "[]"),
        preview_rows=payload[:20],
        created_by=batch.created_by,
        committed_at=batch.committed_at,
        created_at=batch.created_at,
    )


def _normalize_import_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("-", "_")


def _parse_non_negative_number(value, cast, field_name: str):
    if value in (None, ""):
        return cast(0)
    number = cast(value)
    if number < 0:
        raise ValueError(f"{field_name} cannot be negative")
    return number


@router.post("/auth/login", response_model=TokenResponse)
def login(
    form: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db),
    x_device_id: str | None = Header(default=None, alias="X-Device-Id"),
    x_device_token: str | None = Header(default=None, alias="X-Device-Token"),
    x_device_name: str | None = Header(default=None, alias="X-Device-Name"),
):
    user = db.scalar(select(User).where(func.lower(User.email) == form.username.strip().lower()))
    organization = db.get(Organization, user.organization_id) if user else None
    if (
        not user
        or not user.is_active
        or (not user.is_platform_admin and (not organization or not organization.is_active))
        or not verify_password(form.password, user.password_hash)
    ):
        raise HTTPException(
            status_code=401,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    device_id = None
    if x_device_id or x_device_token:
        if not x_device_id or not x_device_token or not (16 <= len(x_device_id) <= 128) or len(x_device_token) < 32:
            raise HTTPException(status_code=400, detail="Valid device id and device token are required together")
        if any(character not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_" for character in x_device_id):
            raise HTTPException(status_code=400, detail="Device id contains invalid characters")
        token_hash = sha256(x_device_token.encode("utf-8")).hexdigest()
        device = db.scalar(select(UserDevice).where(
            UserDevice.organization_id == user.organization_id, UserDevice.device_id == x_device_id
        ))
        if device and device.user_id != user.id:
            raise HTTPException(status_code=409, detail="This device is bound to another account")
        if device and (not device.is_active or device.revoked_at is not None):
            raise HTTPException(status_code=401, detail="This device registration has been revoked")
        if device and not secrets.compare_digest(device.device_token_hash, token_hash):
            raise HTTPException(status_code=401, detail="Device authentication failed")
        if not device:
            device = UserDevice(
                organization_id=user.organization_id,
                user_id=user.id,
                device_id=x_device_id,
                device_token_hash=token_hash,
                device_name=(x_device_name or "Registered device")[:255],
            )
            db.add(device)
        else:
            device.device_name = (x_device_name or device.device_name or "Registered device")[:255]
            device.last_seen_at = datetime.utcnow()
        db.commit()
        device_id = x_device_id
    token, expires_in = create_access_token(user.id, user.organization_id, device_id=device_id)
    return TokenResponse(access_token=token, expires_in=expires_in, user=user, device_id=device_id)


@router.get("/auth/me", response_model=UserRead)
def auth_me(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    if actor.user_id is None:
        raise HTTPException(status_code=401, detail="Authenticated user required")
    user = db.get(User, actor.user_id)
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


@router.post("/users/invitations", response_model=InvitationCreated)
def create_user_invitation(
    payload: InvitationCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN)
    email = payload.email.strip().lower()
    if db.scalar(select(User.id).where(func.lower(User.email) == email)):
        raise HTTPException(status_code=409, detail="A user with this email already exists")
    now = datetime.utcnow()
    pending = db.scalars(
        select(UserInvitation).where(
            func.lower(UserInvitation.email) == email,
            UserInvitation.used_at.is_(None),
        )
    ).all()
    for invitation in pending:
        invitation.used_at = now
    raw_token = secrets.token_urlsafe(32)
    invitation = UserInvitation(
        email=email,
        name=payload.name.strip(),
        role=payload.role,
        token_hash=sha256(raw_token.encode()).hexdigest(),
        invited_by=actor.user_id,
        expires_at=now + timedelta(hours=settings.invitation_expire_hours),
    )
    db.add(invitation)
    db.commit()
    db.refresh(invitation)
    base_url = settings.frontend_public_url.rstrip("/")
    return InvitationCreated(
        id=invitation.id,
        email=invitation.email,
        name=invitation.name,
        role=invitation.role,
        expires_at=invitation.expires_at,
        invitation_url=f"{base_url}/accept-invitation?token={raw_token}",
    )


def _valid_invitation(db: Session, raw_token: str) -> UserInvitation:
    token_hash = sha256(raw_token.encode()).hexdigest()
    invitation = db.scalar(select(UserInvitation).where(UserInvitation.token_hash == token_hash))
    if not invitation or invitation.used_at is not None or invitation.expires_at <= datetime.utcnow():
        raise HTTPException(status_code=400, detail="Invitation is invalid or expired")
    organization = db.get(Organization, invitation.organization_id)
    if not organization or not organization.is_active:
        raise HTTPException(status_code=400, detail="Organization is inactive")
    return invitation


@router.get("/auth/invitations/{token}", response_model=InvitationInfo)
def invitation_info(token: str, db: Session = Depends(get_db)):
    invitation = _valid_invitation(db, token)
    organization = db.get(Organization, invitation.organization_id)
    return InvitationInfo(
        email=invitation.email,
        name=invitation.name,
        role=invitation.role,
        organization_name=organization.name,
        expires_at=invitation.expires_at,
    )


@router.post("/auth/invitations/accept", response_model=UserRead)
def accept_invitation(payload: InvitationAccept, db: Session = Depends(get_db)):
    invitation = _valid_invitation(db, payload.token)
    if db.scalar(select(User.id).where(func.lower(User.email) == invitation.email.lower())):
        raise HTTPException(status_code=409, detail="A user with this email already exists")
    user = User(
        organization_id=invitation.organization_id,
        name=invitation.name,
        email=invitation.email,
        role=invitation.role,
        password_hash=hash_password(payload.password),
        is_active=True,
    )
    invitation.used_at = datetime.utcnow()
    db.add_all([user, invitation])
    db.commit()
    db.refresh(user)
    return user


def _organization_read(db: Session, organization: Organization) -> OrganizationRead:
    return OrganizationRead(
        id=organization.id,
        name=organization.name,
        slug=organization.slug,
        is_active=organization.is_active,
        total_users=db.scalar(select(func.count(User.id)).where(User.organization_id == organization.id)) or 0,
        total_parts=db.scalar(select(func.count(Part.id)).where(Part.organization_id == organization.id)) or 0,
        total_work_orders=db.scalar(
            select(func.count(WorkOrder.id)).where(WorkOrder.organization_id == organization.id)
        ) or 0,
        created_at=organization.created_at,
    )


@router.get("/platform/organizations", response_model=list[OrganizationRead])
def list_organizations(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_platform_admin(actor)
    db.info.pop("organization_id", None)
    organizations = db.scalars(select(Organization).order_by(Organization.id.asc())).all()
    return [_organization_read(db, organization) for organization in organizations]


@router.post("/platform/organizations", response_model=OrganizationRead)
def create_organization(
    payload: OrganizationCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_platform_admin(actor)
    db.info.pop("organization_id", None)
    slug = payload.slug.strip().lower()
    email = payload.admin_email.strip().lower()
    if db.scalar(select(Organization.id).where(Organization.slug == slug)):
        raise HTTPException(status_code=409, detail="Organization slug already exists")
    if db.scalar(select(User.id).where(func.lower(User.email) == email)):
        raise HTTPException(status_code=409, detail="Administrator email already exists")

    organization = Organization(name=payload.name.strip(), slug=slug)
    db.add(organization)
    db.flush()
    administrator = User(
        organization_id=organization.id,
        name=payload.admin_name.strip(),
        email=email,
        role=UserRole.ADMIN,
        password_hash=hash_password(payload.admin_password),
        is_active=True,
        is_platform_admin=False,
    )
    db.add(administrator)
    db.commit()
    db.refresh(organization)
    return _organization_read(db, organization)


@router.patch("/platform/organizations/{organization_id}", response_model=OrganizationRead)
def update_organization(
    organization_id: int,
    payload: OrganizationUpdate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_platform_admin(actor)
    db.info.pop("organization_id", None)
    organization = db.get(Organization, organization_id)
    if not organization:
        raise HTTPException(status_code=404, detail="Organization not found")
    organization.is_active = payload.is_active
    db.add(organization)
    db.commit()
    db.refresh(organization)
    return _organization_read(db, organization)


def _require_tenant_user(db: Session, user_id: int | None, field_name: str) -> None:
    if user_id is not None and not db.get(User, user_id):
        raise HTTPException(status_code=400, detail=f"{field_name} must reference a user in the current organization")


def _audit(
    db: Session,
    actor: Actor,
    action: str,
    entity_type: str,
    entity_id: int | None = None,
    metadata: dict | None = None,
):
    audit_metadata = {
        "actor_role": actor.role.value,
        "auth_method": actor.auth_method,
        "device_id": actor.device_id,
        "device_record_id": actor.device_record_id,
        "claim_version": actor.claim_version,
        **(metadata or {}),
    }
    db.add(
        AuditLog(
            user_id=actor.user_id,
            action=action,
            entity_type=entity_type,
            entity_id=entity_id,
            metadata_json=json.dumps(audit_metadata, default=str, separators=(",", ":")),
            timestamp=datetime.utcnow(),
        )
    )


def _require_account_reauthentication(db: Session, actor: Actor, password: str | None) -> None:
    if actor.auth_method == "test":
        return
    if actor.auth_method != "bearer" or actor.user_id is None:
        raise HTTPException(status_code=401, detail="Bearer authentication required")
    user = db.get(User, actor.user_id)
    if not password or not user or not verify_password(password, user.password_hash):
        raise HTTPException(status_code=401, detail="Account password verification failed")


def _work_order_read_for_actor(db: Session, actor: Actor, item: WorkOrder) -> WorkOrderRead:
    payload = WorkOrderRead.model_validate(item).model_dump()
    claimant = db.get(User, item.claimed_by_id) if item.claimed_by_id else None
    completed_by = db.get(User, item.completed_by_id) if item.completed_by_id else None
    completed_device = db.get(UserDevice, item.completed_device_id) if item.completed_device_id else None
    is_test_actor = actor.auth_method == "test"
    is_engineer_owner = bool(
        actor.role == UserRole.ENGINEER
        and actor.user_id == item.claimed_by_id
        and actor.device_verified
        and actor.device_record_id == item.claimed_device_id
    )
    execution_open = not item.is_locked and item.status.upper() != "PENDING_APPROVAL"
    payload.update(
        claimed_by_name=claimant.name if claimant else None,
        completed_by_name=completed_by.name if completed_by else None,
        completed_device_name=(
            completed_device.device_name
            if completed_device and (is_engineer_owner or actor.role in {UserRole.ADMIN, UserRole.MANAGER})
            else None
        ),
        can_claim=bool(
            actor.role == UserRole.ENGINEER
            and actor.device_verified
            and item.claimed_by_id is None
            and execution_open
            and item.status.upper() != "COMPLETED"
        ),
        can_edit=bool((is_engineer_owner or actor.role == UserRole.ADMIN or is_test_actor) and execution_open),
        can_complete=bool((is_engineer_owner or is_test_actor) and execution_open),
    )
    if actor.role not in {UserRole.ADMIN, UserRole.MANAGER} and not is_engineer_owner and not is_test_actor:
        payload["revenue"] = 0.0
        payload["labor_cost"] = 0.0
        payload["customer_signature_name"] = None
        payload["customer_signature_data"] = None
        payload["customer_signed_at"] = None
        payload["claimed_device_id"] = None
        payload["completed_device_id"] = None
    return WorkOrderRead(**payload)


def _image_extension(data: bytes) -> str | None:
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith((b"GIF87a", b"GIF89a")):
        return ".gif"
    if len(data) >= 12 and data.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp"
    if len(data) >= 12 and data[4:8] == b"ftyp" and data[8:12] in {
        b"heic", b"heix", b"hevc", b"hevx", b"mif1",
    }:
        return ".heic"
    return None


@router.post("/uploads/work-order-parts")
async def upload_work_order_part_photo(
    work_order_id: int = Form(..., ge=1),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    work_order = require_work_order_execution_scope(db, actor, work_order_id)
    if not work_order or work_order.is_locked or work_order.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=409, detail="Work order cannot accept uploads in its current state")

    data = await file.read(settings.max_image_upload_bytes + 1)
    if len(data) > settings.max_image_upload_bytes:
        raise HTTPException(status_code=413, detail="Image exceeds the configured upload limit")
    ext = _image_extension(data)
    if not ext:
        raise HTTPException(status_code=400, detail="Unsupported or invalid image file")

    target_dir = Path("uploads/work-order-parts")
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid4().hex}{ext}"
    target_path = target_dir / filename

    target_path.write_bytes(data)
    _audit(
        db,
        actor,
        "upload_work_order_photo",
        "work_order",
        work_order_id,
        {"url": f"/uploads/work-order-parts/{filename}"},
    )
    db.commit()
    return {"url": f"/uploads/work-order-parts/{filename}"}


def _audio_extension(data: bytes) -> tuple[str, str] | None:
    if data.startswith(b"\x1aE\xdf\xa3"):
        return ".webm", "audio/webm"
    if data.startswith(b"OggS"):
        return ".ogg", "audio/ogg"
    if len(data) >= 12 and data.startswith(b"RIFF") and data[8:12] == b"WAVE":
        return ".wav", "audio/wav"
    if len(data) >= 12 and data[4:8] == b"ftyp":
        return ".m4a", "audio/mp4"
    if data.startswith(b"ID3") or (len(data) >= 2 and data[0] == 0xFF and data[1] & 0xE0 == 0xE0):
        return ".mp3", "audio/mpeg"
    return None


@router.post("/work-orders/{work_order_id}/voice-notes", response_model=WorkOrderVoiceNoteRead)
async def create_work_order_voice_note(
    work_order_id: int,
    file: UploadFile = File(...),
    duration_seconds: float | None = Form(None, ge=0, le=7200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    work_order = require_work_order_execution_scope(db, actor, work_order_id)
    if not work_order or work_order.is_locked or work_order.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=400, detail="Voice notes cannot be added to this work order")
    data = await file.read(settings.max_audio_upload_bytes + 1)
    if len(data) > settings.max_audio_upload_bytes:
        raise HTTPException(status_code=413, detail="Audio exceeds the configured upload limit")
    audio_type = _audio_extension(data)
    if not audio_type:
        raise HTTPException(status_code=400, detail="Unsupported or invalid audio file")
    extension, mime_type = audio_type
    target_dir = Path("uploads/work-order-voice-notes")
    target_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid4().hex}{extension}"
    (target_dir / filename).write_bytes(data)
    note = WorkOrderVoiceNote(
        organization_id=actor.organization_id,
        work_order_id=work_order_id,
        created_by=actor.user_id,
        audio_url=f"/uploads/work-order-voice-notes/{filename}",
        mime_type=mime_type,
        duration_seconds=duration_seconds,
    )
    db.add(note)
    _audit(db, actor, "create_voice_note", "work_order_voice_note", None, {"work_order_id": work_order_id})
    db.commit()
    db.refresh(note)
    return note


@router.get("/work-orders/{work_order_id}/voice-notes", response_model=list[WorkOrderVoiceNoteRead])
def list_work_order_voice_notes(
    work_order_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_work_order_scope(db, actor, work_order_id)
    return db.scalars(
        select(WorkOrderVoiceNote)
        .where(
            WorkOrderVoiceNote.organization_id == actor.organization_id,
            WorkOrderVoiceNote.work_order_id == work_order_id,
        )
        .order_by(WorkOrderVoiceNote.id.desc())
    ).all()


@router.post("/users", response_model=UserRead)
def create_user(payload: UserCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN)
    data = payload.model_dump(exclude={"password"})
    item = User(**data, password_hash=hash_password(payload.password) if payload.password else None)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.post("/users/{user_id}/set-password", status_code=204)
def set_user_password(
    user_id: int,
    payload: PasswordSet,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.password_hash = hash_password(payload.password)
    db.add(user)
    db.commit()


@router.get("/users", response_model=list[UserRead])
def list_users(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    return db.scalars(select(User).order_by(User.id.desc()).offset(skip).limit(limit)).all()


@router.post("/warehouses", response_model=WarehouseRead)
def create_warehouse(payload: WarehouseCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    _require_tenant_user(db, payload.assigned_user_id, "assigned_user_id")
    values = payload.model_dump()
    values["warehouse_type"] = (payload.warehouse_type or "main").strip().lower()
    owner = db.get(User, payload.assigned_user_id) if payload.assigned_user_id else None
    if owner and owner.role == UserRole.ENGINEER and values["warehouse_type"] == "main":
        values["warehouse_type"] = "van"
    if values["warehouse_type"] not in {"main", "van"}:
        raise HTTPException(status_code=422, detail="warehouse_type must be main or van")
    if values["warehouse_type"] == "van" and (
        not owner or not owner.is_active or owner.role != UserRole.ENGINEER
    ):
        raise HTTPException(status_code=422, detail="A van warehouse must be assigned to an active engineer")
    values["code"] = (payload.code or payload.name).strip().upper().replace(" ", "-")
    item = Warehouse(**values)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/warehouses", response_model=list[WarehouseRead])
def list_warehouses(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    return db.scalars(select(Warehouse).order_by(Warehouse.id.desc()).offset(skip).limit(limit)).all()


@router.post("/storage-locations", response_model=StorageLocationRead)
def create_storage_location(
    payload: StorageLocationCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    if not db.get(Warehouse, payload.warehouse_id):
        raise HTTPException(status_code=404, detail="Warehouse not found")
    item = StorageLocation(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/storage-locations", response_model=list[StorageLocationRead])
def list_storage_locations(
    warehouse_id: int | None = None,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    query = select(StorageLocation).order_by(StorageLocation.code.asc())
    if warehouse_id is not None:
        if not db.get(Warehouse, warehouse_id):
            raise HTTPException(status_code=404, detail="Warehouse not found")
        query = query.where(StorageLocation.warehouse_id == warehouse_id)
    return db.scalars(query).all()


@router.post("/parts", response_model=PartRead)
def create_part(payload: PartCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    item = Part(**payload.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.post("/parts/recognition/observations", response_model=PartMachineAssociationRead)
async def record_part_observation(
    machine_model: str = Form(..., min_length=1),
    part_id: int | None = Form(None),
    part_number: str | None = Form(None),
    part_name: str | None = Form(None),
    file: UploadFile | None = File(None),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    part = db.get(Part, part_id) if part_id else None
    if not part and part_number:
        part = db.scalar(select(Part).where(Part.part_number == part_number.strip()))
    if not part:
        if not part_number or not part_name:
            raise HTTPException(status_code=400, detail="part_id or part_number and part_name are required")
        part = Part(part_number=part_number.strip(), name=part_name.strip(), machine_type=machine_model.strip())
        db.add(part)
        db.flush()
    photo_url = None
    if file:
        data = await file.read(settings.max_image_upload_bytes + 1)
        if len(data) > settings.max_image_upload_bytes:
            raise HTTPException(status_code=413, detail="Image exceeds the configured upload limit")
        ext = _image_extension(data)
        if not ext:
            raise HTTPException(status_code=400, detail="Unsupported or invalid image file")
        target_dir = Path("uploads/part-observations")
        target_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{uuid4().hex}{ext}"
        (target_dir / filename).write_bytes(data)
        photo_url = f"/uploads/part-observations/{filename}"
    association = db.scalar(select(PartMachineAssociation).where(
        PartMachineAssociation.machine_model == machine_model.strip(), PartMachineAssociation.part_id == part.id
    ))
    if association:
        association.confirmed_count += 1
        association.last_confirmed_at = datetime.utcnow()
        association.photo_url = photo_url or association.photo_url
    else:
        association = PartMachineAssociation(machine_model=machine_model.strip(), part_id=part.id, photo_url=photo_url)
        db.add(association)
    _audit(db, actor, "record_part_observation", "part_machine_association", association.id, {"machine_model": machine_model})
    db.commit()
    db.refresh(association)
    return association


@router.get("/parts/recognition/suggestions", response_model=list[PartMachineAssociationRead])
def part_recognition_suggestions(machine_model: str, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    return db.scalars(select(PartMachineAssociation).where(PartMachineAssociation.machine_model.ilike(f"%{machine_model.strip()}%")).order_by(PartMachineAssociation.confirmed_count.desc())).all()


@router.get("/parts", response_model=list[PartRead])
def list_parts(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    return db.scalars(select(Part).order_by(Part.id.desc()).offset(skip).limit(limit)).all()


@router.post("/work-orders", response_model=WorkOrderRead)
def create_work_order(payload: WorkOrderCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    if payload.status.strip().upper() in {"COMPLETED", "PENDING_APPROVAL", "APPROVAL_REJECTED"}:
        raise HTTPException(status_code=400, detail="Terminal work order status requires the completion workflow")
    _require_tenant_user(db, payload.assigned_user_id, "assigned_user_id")
    _require_tenant_user(db, payload.engineer_id, "engineer_id")
    _require_tenant_user(db, payload.assistant_id, "assistant_id")
    data = payload.model_dump()
    customer, equipment = _require_tenant_service_links(db, actor, data.get("customer_id"), data.get("equipment_id"))
    if equipment and not data.get("customer_id"):
        data["customer_id"] = equipment.customer_id
        customer = db.get(Customer, equipment.customer_id) if equipment.customer_id else None
    if customer:
        data["store_name"] = data.get("store_name") or customer.name
        data["outlet_name"] = data.get("outlet_name") or customer.name
        data["contact_phone"] = data.get("contact_phone") or customer.phone
        data["address"] = data.get("address") or customer.address
        data["city"] = data.get("city") or customer.city
        data["state"] = data.get("state") or customer.state
        data["zip"] = data.get("zip") or customer.zip
    if equipment:
        data["machine_type"] = data.get("machine_type") or equipment.model
    if not data.get("ticket_number"):
        data["ticket_number"] = data.get("wo_number")
    if not data.get("wo_number"):
        data["wo_number"] = data.get("ticket_number")
    if not data.get("store_name") and data.get("outlet_name"):
        data["store_name"] = data["outlet_name"]
    if not data.get("outlet_name") and data.get("store_name"):
        data["outlet_name"] = data["store_name"]
    if not data.get("problem_description") and data.get("description"):
        data["problem_description"] = data["description"]
    if not data.get("description") and data.get("problem_description"):
        data["description"] = data["problem_description"]
    if not data.get("engineer_id") and data.get("assigned_user_id"):
        data["engineer_id"] = data["assigned_user_id"]
    if not data.get("assigned_user_id") and data.get("engineer_id"):
        data["assigned_user_id"] = data["engineer_id"]
    item = WorkOrder(**data)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/work-orders", response_model=list[WorkOrderRead])
def list_work_orders(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    technician_id: int | None = Query(default=None, ge=1),
    status: str | None = Query(default=None),
    city: str | None = Query(default=None),
    job_type: str | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    scope: str = Query(default="all", pattern="^(all|mine|available)$"),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    stmt = select(WorkOrder).order_by(WorkOrder.id.desc())
    if actor.role == UserRole.ENGINEER and actor.user_id:
        if scope == "mine":
            stmt = stmt.where(WorkOrder.claimed_by_id == actor.user_id)
        elif scope == "available":
            stmt = stmt.where(WorkOrder.claimed_by_id.is_(None), WorkOrder.is_locked.is_(False))
    elif actor.role not in {UserRole.ADMIN, UserRole.MANAGER}:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    if technician_id:
        stmt = stmt.where((WorkOrder.assigned_user_id == technician_id) | (WorkOrder.engineer_id == technician_id))
    if status:
        stmt = stmt.where(func.lower(WorkOrder.status) == status.lower())
    if city:
        stmt = stmt.where(func.lower(WorkOrder.city) == city.lower())
    if job_type:
        stmt = stmt.where(func.lower(WorkOrder.job_type) == job_type.lower())
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            (func.lower(WorkOrder.ticket_number).like(like))
            | (func.lower(func.coalesce(WorkOrder.wo_number, "")).like(like))
            | (func.lower(func.coalesce(WorkOrder.outlet_name, "")).like(like))
            | (func.lower(func.coalesce(WorkOrder.address, "")).like(like))
        )
    if date_from:
        stmt = stmt.where(WorkOrder.schedule_date >= date_from)
    if date_to:
        stmt = stmt.where(WorkOrder.schedule_date <= date_to)
    rows = db.scalars(stmt.offset(skip).limit(limit)).all()
    return [_work_order_read_for_actor(db, actor, item) for item in rows]


@router.post("/work-orders/{work_order_id}/claim", response_model=WorkOrderRead)
def claim_work_order(
    work_order_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    if actor.role != UserRole.ENGINEER:
        raise HTTPException(status_code=403, detail="Only engineers can claim work orders")
    require_bound_device(actor)
    require_work_order_scope(db, actor, work_order_id)
    item = db.get(WorkOrder, work_order_id)
    if not item:
        raise HTTPException(status_code=404, detail="Work order not found")
    if item.is_locked or item.status.upper() in {"COMPLETED", "PENDING_APPROVAL"}:
        raise HTTPException(status_code=409, detail="Work order cannot be claimed in its current state")
    if item.claimed_by_id is not None:
        if item.claimed_by_id == actor.user_id and item.claimed_device_id == actor.device_record_id:
            return _work_order_read_for_actor(db, actor, item)
        if item.claimed_by_id == actor.user_id:
            raise HTTPException(status_code=409, detail="Work order is already bound to another registered device")
        raise HTTPException(status_code=409, detail="Work order has already been claimed")

    result = db.execute(
        update(WorkOrder)
        .where(
            WorkOrder.id == work_order_id,
            WorkOrder.organization_id == actor.organization_id,
            WorkOrder.claimed_by_id.is_(None),
            WorkOrder.is_locked.is_(False),
            func.upper(WorkOrder.status).notin_({"COMPLETED", "PENDING_APPROVAL"}),
        )
        .values(
            claimed_by_id=actor.user_id,
            claimed_device_id=actor.device_record_id,
            claimed_at=datetime.utcnow(),
            claim_version=WorkOrder.claim_version + 1,
            assigned_user_id=actor.user_id,
            engineer_id=actor.user_id,
        )
        .execution_options(synchronize_session=False)
    )
    if result.rowcount != 1:
        db.rollback()
        raise HTTPException(status_code=409, detail="Work order was claimed by another engineer")
    db.refresh(item)
    _audit(
        db,
        actor,
        "claim_work_order",
        "work_order",
        work_order_id,
        {"claimed_by_id": actor.user_id, "claimed_device_id": actor.device_record_id, "new_claim_version": item.claim_version},
    )
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.post("/work-orders/{work_order_id}/release", response_model=WorkOrderRead)
def release_work_order_claim(
    work_order_id: int,
    payload: WorkOrderClaimRelease,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    require_work_order_scope(db, actor, work_order_id)
    item = db.get(WorkOrder, work_order_id)
    if not item or item.claimed_by_id is None:
        raise HTTPException(status_code=409, detail="Work order is not currently claimed")
    if item.is_locked or item.status.upper() in {"COMPLETED", "PENDING_APPROVAL"}:
        raise HTTPException(status_code=409, detail="Claim cannot be released in the current state")
    previous_user_id = item.claimed_by_id
    previous_device_id = item.claimed_device_id
    if item.assigned_user_id == previous_user_id:
        item.assigned_user_id = None
    if item.engineer_id == previous_user_id:
        item.engineer_id = None
    item.claimed_by_id = None
    item.claimed_device_id = None
    item.claimed_at = None
    item.claim_version += 1
    _audit(
        db,
        actor,
        "release_work_order_claim",
        "work_order",
        work_order_id,
        {
            "reason": payload.reason.strip(),
            "previous_claimed_by_id": previous_user_id,
            "previous_claimed_device_id": previous_device_id,
            "new_claim_version": item.claim_version,
        },
    )
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.get("/work-orders/{work_order_id}/service-context", response_model=WorkOrderServiceContext)
def get_work_order_service_context(
    work_order_id: int,
    history_limit: int = Query(default=5, ge=1, le=20),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_work_order_scope(db, actor, work_order_id)
    current = db.get(WorkOrder, work_order_id)
    if not current:
        raise HTTPException(status_code=404, detail="Work order not found")
    customer = db.get(Customer, current.customer_id) if current.customer_id else None
    equipment = db.get(Equipment, current.equipment_id) if current.equipment_id else None

    history_stmt = select(WorkOrder).where(
        WorkOrder.organization_id == actor.organization_id,
        WorkOrder.id != current.id,
        func.lower(WorkOrder.status) == "completed",
    )
    if current.equipment_id:
        history_stmt = history_stmt.where(WorkOrder.equipment_id == current.equipment_id)
    elif current.customer_id and current.machine_type:
        history_stmt = history_stmt.where(
            WorkOrder.customer_id == current.customer_id,
            func.lower(func.coalesce(WorkOrder.machine_type, "")) == current.machine_type.lower(),
        )
    elif current.customer_id:
        history_stmt = history_stmt.where(WorkOrder.customer_id == current.customer_id)
    else:
        site_name = (current.outlet_name or current.store_name or "").strip().lower()
        machine = (current.machine_type or "").strip().lower()
        if not site_name or not machine:
            history_rows = []
            history_stmt = None
        else:
            history_stmt = history_stmt.where(
                func.lower(func.coalesce(WorkOrder.machine_type, "")) == machine,
                or_(
                    func.lower(func.coalesce(WorkOrder.outlet_name, "")) == site_name,
                    func.lower(func.coalesce(WorkOrder.store_name, "")) == site_name,
                ),
            )
    if history_stmt is not None:
        history_rows = db.scalars(
            history_stmt.order_by(WorkOrder.completed_at.desc(), WorkOrder.id.desc()).limit(history_limit)
        ).all()

    history = []
    for row in history_rows:
        part_rows = db.execute(
            select(Part.part_number, Part.name, func.sum(WorkOrderPart.quantity))
            .join(WorkOrderPart, WorkOrderPart.part_id == Part.id)
            .where(
                WorkOrderPart.organization_id == actor.organization_id,
                WorkOrderPart.work_order_id == row.id,
            )
            .group_by(Part.part_number, Part.name)
            .order_by(Part.part_number)
        ).all()
        history.append({
            "id": row.id,
            "ticket_number": row.ticket_number,
            "schedule_date": row.schedule_date,
            "job_type": row.job_type,
            "problem_description": row.problem_description,
            "repair_result": row.repair_result,
            "status": row.status,
            "completed_at": row.completed_at,
            "engineer_id": row.engineer_id,
            "parts_used": [
                {"part_number": part_number, "name": name, "quantity": int(quantity or 0)}
                for part_number, name, quantity in part_rows
            ],
        })
    return {
        "customer": customer,
        "equipment": equipment,
        "fallback_customer_name": current.outlet_name or current.store_name,
        "fallback_contact_phone": current.contact_phone,
        "fallback_equipment_model": current.machine_type,
        "history": history,
    }


@router.patch("/work-orders/{work_order_id}", response_model=WorkOrderRead)
def update_work_order(
    work_order_id: int,
    payload: WorkOrderUpdate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    item = require_work_order_write_scope(db, actor, work_order_id)
    if item.is_locked:
        raise HTTPException(status_code=400, detail="Work order is locked and cannot be edited")
    if item.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=409, detail="Pending completion evidence is frozen until approval or rejection")

    updates = payload.model_dump(exclude_unset=True)
    if actor.role == UserRole.ENGINEER:
        blocked = {"revenue", "labor_cost", "assigned_user_id", "engineer_id", "assistant_id", "customer_id", "equipment_id", "status"}
        for key in blocked:
            updates.pop(key, None)
    if str(updates.get("status", "")).strip().upper() in {"COMPLETED", "PENDING_APPROVAL", "APPROVAL_REJECTED"}:
        raise HTTPException(status_code=400, detail="Terminal work order status requires the completion workflow")
    if "ticket_number" in updates and "wo_number" not in updates:
        updates["wo_number"] = updates["ticket_number"]
    if "wo_number" in updates and "ticket_number" not in updates:
        updates["ticket_number"] = updates["wo_number"]
    if "outlet_name" in updates and "store_name" not in updates:
        updates["store_name"] = updates["outlet_name"]
    if "store_name" in updates and "outlet_name" not in updates:
        updates["outlet_name"] = updates["store_name"]
    if "description" in updates and "problem_description" not in updates:
        updates["problem_description"] = updates["description"]
    if "problem_description" in updates and "description" not in updates:
        updates["description"] = updates["problem_description"]
    if "engineer_id" in updates and "assigned_user_id" not in updates:
        updates["assigned_user_id"] = updates["engineer_id"]
    if "assigned_user_id" in updates and "engineer_id" not in updates:
        updates["engineer_id"] = updates["assigned_user_id"]

    _require_tenant_user(db, updates.get("assigned_user_id"), "assigned_user_id")
    _require_tenant_user(db, updates.get("engineer_id"), "engineer_id")
    _require_tenant_user(db, updates.get("assistant_id"), "assistant_id")
    target_customer_id = updates.get("customer_id", item.customer_id)
    target_equipment_id = updates.get("equipment_id", item.equipment_id)
    if item.claimed_by_id is not None and any(
        field in updates and updates[field] != getattr(item, field)
        for field in {"assigned_user_id", "engineer_id"}
    ):
        raise HTTPException(status_code=409, detail="Release the authenticated claim before reassigning this work order")
    customer, equipment = _require_tenant_service_links(db, actor, target_customer_id, target_equipment_id)
    if equipment and target_customer_id is None:
        updates["customer_id"] = equipment.customer_id
    if customer and equipment and equipment.customer_id not in {None, customer.id}:
        raise HTTPException(status_code=400, detail="Equipment does not belong to the selected customer")

    for key, value in updates.items():
        setattr(item, key, value)

    db.add(item)
    _audit(db, actor, "update_work_order", "work_order", work_order_id, {"changed_fields": sorted(updates)})
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.post("/work-orders/{work_order_id}/start", response_model=WorkOrderRead)
def start_work_order(
    work_order_id: int,
    payload: WorkOrderFlowAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    item = require_work_order_execution_scope(db, actor, work_order_id)
    if item.is_locked:
        raise HTTPException(status_code=400, detail="Work order is locked and cannot be started")
    if item.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=409, detail="Work order is awaiting manager approval")
    item.status = "IN_PROGRESS"
    item.started_at = item.started_at or datetime.utcnow()
    db.add(item)
    db.add(JobStatus(work_order_id=work_order_id, status="IN_PROGRESS", timestamp=datetime.utcnow()))
    _audit(db, actor, "start_job", "work_order", work_order_id)
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.post("/work-orders/{work_order_id}/complete", response_model=WorkOrderRead)
def complete_work_order(
    work_order_id: int,
    payload: WorkOrderFlowAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    item = require_work_order_owner_scope(db, actor, work_order_id)
    _require_account_reauthentication(db, actor, payload.account_password)
    if item.is_locked:
        raise HTTPException(status_code=400, detail="Work order already completed")
    _apply_completion_payload(item, payload)
    policy = _effective_completion_policy(db, actor.organization_id, item.job_type)
    _validate_completion_evidence(db, item, policy)
    if policy.get("require_manager_approval") and actor.role not in {UserRole.ADMIN, UserRole.MANAGER}:
        item.status = "PENDING_APPROVAL"
        item.completion_requested_by = actor.user_id
        item.completion_requested_at = datetime.utcnow()
        db.add(item)
        db.add(JobStatus(work_order_id=work_order_id, status="PENDING_APPROVAL", timestamp=datetime.utcnow()))
        _audit(db, actor, "request_completion", "work_order", work_order_id, {"policy": policy})
        db.commit()
        db.refresh(item)
        return _work_order_read_for_actor(db, actor, item)
    if policy.get("require_manager_approval"):
        item.completion_approved_by = actor.user_id
        item.completion_approved_at = datetime.utcnow()
    return _finalize_work_order(db, actor, item)


def _normalize_job_type(job_type: str | None) -> str:
    return (job_type or "*").strip().lower() or "*"


def _policy_dict(policy: CompletionPolicy | None, organization_id: int, source: str) -> dict:
    if not policy:
        return {
            "id": None, "organization_id": organization_id, "job_type": None, "source": "legacy_default",
            "require_repair_result": False, "require_customer_signature": False,
            "require_completion_photo": False, "require_all_checklist_items": False,
            "require_parts_usage": False, "require_manager_approval": False,
            "created_at": None, "updated_at": None,
        }
    return {
        "id": policy.id, "organization_id": policy.organization_id,
        "job_type": None if policy.job_type_key == "*" else policy.job_type_key, "source": source,
        "require_repair_result": policy.require_repair_result,
        "require_customer_signature": policy.require_customer_signature,
        "require_completion_photo": policy.require_completion_photo,
        "require_all_checklist_items": policy.require_all_checklist_items,
        "require_parts_usage": policy.require_parts_usage,
        "require_manager_approval": policy.require_manager_approval,
        "created_at": policy.created_at, "updated_at": policy.updated_at,
    }


def _effective_completion_policy(db: Session, organization_id: int, job_type: str | None) -> dict:
    key = _normalize_job_type(job_type)
    policy = None
    source = "organization_default"
    if key != "*":
        policy = db.scalar(select(CompletionPolicy).where(
            CompletionPolicy.organization_id == organization_id, CompletionPolicy.job_type_key == key
        ))
        if policy:
            source = "job_type"
    if not policy:
        policy = db.scalar(select(CompletionPolicy).where(
            CompletionPolicy.organization_id == organization_id, CompletionPolicy.job_type_key == "*"
        ))
    return _policy_dict(policy, organization_id, source)


def _apply_completion_payload(item: WorkOrder, payload: WorkOrderFlowAction) -> None:
    if payload.repair_result is not None:
        item.repair_result = payload.repair_result.strip() or None
    if payload.checklist_json is not None:
        item.checklist_json = payload.checklist_json
    if payload.customer_signature_name is not None:
        item.customer_signature_name = payload.customer_signature_name.strip() or None
    if payload.customer_signature_data is not None:
        item.customer_signature_data = payload.customer_signature_data
    item.customer_signed_at = datetime.utcnow() if item.customer_signature_name and item.customer_signature_data else None


def _validate_completion_evidence(db: Session, item: WorkOrder, policy: dict) -> None:
    missing: list[str] = []
    if policy.get("require_repair_result") and not (item.repair_result or "").strip():
        missing.append("repair_result")
    if policy.get("require_customer_signature") and not (item.customer_signature_name and item.customer_signature_data):
        missing.append("customer_signature")
    if policy.get("require_all_checklist_items"):
        try:
            checklist = json.loads(item.checklist_json or "")
        except (TypeError, json.JSONDecodeError):
            checklist = None
        required_keys = {"equipment_safe", "site_clean", "customer_briefed"}
        if (
            not isinstance(checklist, dict)
            or not required_keys.issubset(checklist)
            or not all(type(checklist[key]) is bool and checklist[key] for key in required_keys)
        ):
            missing.append("completed_checklist")
    if policy.get("require_completion_photo") and not db.scalar(
        select(QCPicture.id).where(QCPicture.work_order_id == item.id).limit(1)
    ):
        missing.append("completion_photo")
    if policy.get("require_parts_usage") and not db.scalar(
        select(WorkOrderPart.id).where(WorkOrderPart.work_order_id == item.id).limit(1)
    ):
        missing.append("parts_usage")
    if missing:
        raise HTTPException(status_code=422, detail={"message": "Completion evidence is incomplete", "missing": missing})


def _finalize_work_order(db: Session, actor: Actor, item: WorkOrder) -> WorkOrderRead:
    work_order_id = item.id
    completed_by_id = item.claimed_by_id
    completed_device_id = item.claimed_device_id
    if actor.auth_method == "test" and completed_by_id is None:
        completed_by_id = actor.user_id
        completed_device_id = actor.device_record_id
    if actor.auth_method != "test" and (completed_by_id is None or completed_device_id is None):
        raise HTTPException(status_code=409, detail="Work order has no authenticated engineer claim")
    item.completed_by_id = completed_by_id
    item.completed_device_id = completed_device_id
    item.status = "COMPLETED"
    item.completed_at = datetime.utcnow()
    item.is_locked = True
    db.add(item)
    db.add(JobStatus(work_order_id=work_order_id, status="COMPLETED", timestamp=datetime.utcnow()))
    _audit(
        db,
        actor,
        "complete_job",
        "work_order",
        work_order_id,
        {
            "parts_cost": get_work_order_parts_cost(db, work_order_id),
            "completed_by_id": completed_by_id,
            "completed_device_id": completed_device_id,
            "claim_version": item.claim_version,
        },
    )
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


def _require_tenant_service_links(
    db: Session,
    actor: Actor,
    customer_id: int | None,
    equipment_id: int | None,
) -> tuple[Customer | None, Equipment | None]:
    customer = None
    equipment = None
    if customer_id:
        customer = db.scalar(select(Customer).where(Customer.id == customer_id, Customer.organization_id == actor.organization_id))
        if not customer:
            raise HTTPException(status_code=400, detail="customer_id is not available in this organization")
    if equipment_id:
        equipment = db.scalar(select(Equipment).where(Equipment.id == equipment_id, Equipment.organization_id == actor.organization_id))
        if not equipment:
            raise HTTPException(status_code=400, detail="equipment_id is not available in this organization")
    if customer and equipment and equipment.customer_id not in {None, customer.id}:
        raise HTTPException(status_code=400, detail="Equipment does not belong to the selected customer")
    return customer, equipment


@router.post("/customers", response_model=CustomerRead)
def create_customer(payload: CustomerCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    data = payload.model_dump()
    if data.get("account_number") == "":
        data["account_number"] = None
    item = Customer(organization_id=actor.organization_id, **data)
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/completion-policies", response_model=list[CompletionPolicyRead])
def list_completion_policies(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    rows = db.scalars(select(CompletionPolicy).where(
        CompletionPolicy.organization_id == actor.organization_id
    ).order_by(CompletionPolicy.job_type_key, CompletionPolicy.id)).all()
    return [_policy_dict(row, actor.organization_id, "organization_default" if row.job_type_key == "*" else "job_type") for row in rows]


@router.post("/completion-policies", response_model=CompletionPolicyRead)
def upsert_completion_policy(
    payload: CompletionPolicyUpsert,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    key = _normalize_job_type(payload.job_type)
    item = db.scalar(select(CompletionPolicy).where(
        CompletionPolicy.organization_id == actor.organization_id, CompletionPolicy.job_type_key == key
    ))
    values = payload.model_dump(exclude={"job_type"})
    if item:
        for field, value in values.items():
            setattr(item, field, value)
    else:
        item = CompletionPolicy(organization_id=actor.organization_id, job_type_key=key, **values)
    db.add(item)
    db.flush()
    _audit(db, actor, "upsert_completion_policy", "completion_policy", item.id, {"job_type": key, **values})
    db.commit()
    db.refresh(item)
    return _policy_dict(item, actor.organization_id, "organization_default" if key == "*" else "job_type")


@router.get("/work-orders/{work_order_id}/completion-policy", response_model=CompletionPolicyRead)
def get_work_order_completion_policy(
    work_order_id: int, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_work_order_scope(db, actor, work_order_id)
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.ENGINEER)
    item = db.get(WorkOrder, work_order_id)
    if not item:
        raise HTTPException(status_code=404, detail="Work order not found")
    return _effective_completion_policy(db, actor.organization_id, item.job_type)


@router.post("/work-orders/{work_order_id}/request-completion", response_model=WorkOrderRead)
def request_work_order_completion(
    work_order_id: int,
    payload: WorkOrderFlowAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    item = require_work_order_owner_scope(db, actor, work_order_id)
    _require_account_reauthentication(db, actor, payload.account_password)
    if not item or item.is_locked:
        raise HTTPException(status_code=400, detail="Work order cannot request completion")
    if item.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=409, detail="Completion approval is already pending")
    policy = _effective_completion_policy(db, actor.organization_id, item.job_type)
    if not policy.get("require_manager_approval"):
        raise HTTPException(status_code=409, detail="This work order does not require manager approval")
    _apply_completion_payload(item, payload)
    _validate_completion_evidence(db, item, policy)
    item.status = "PENDING_APPROVAL"
    item.completion_requested_by = actor.user_id
    item.completion_requested_at = datetime.utcnow()
    db.add(item)
    db.add(JobStatus(work_order_id=work_order_id, status="PENDING_APPROVAL", timestamp=datetime.utcnow()))
    _audit(db, actor, "request_completion", "work_order", work_order_id)
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.post("/work-orders/{work_order_id}/approve-completion", response_model=WorkOrderRead)
def approve_work_order_completion(
    work_order_id: int, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_work_order_scope(db, actor, work_order_id)
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    item = db.get(WorkOrder, work_order_id)
    if not item or item.status != "PENDING_APPROVAL" or item.is_locked:
        raise HTTPException(status_code=409, detail="Work order is not pending completion approval")
    policy = _effective_completion_policy(db, actor.organization_id, item.job_type)
    _validate_completion_evidence(db, item, policy)
    item.completion_approved_by = actor.user_id
    item.completion_approved_at = datetime.utcnow()
    _audit(db, actor, "approve_completion", "work_order", work_order_id)
    return _finalize_work_order(db, actor, item)


@router.post("/work-orders/{work_order_id}/reject-completion", response_model=WorkOrderRead)
def reject_work_order_completion(
    work_order_id: int,
    payload: WorkOrderFlowAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_work_order_scope(db, actor, work_order_id)
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    item = db.get(WorkOrder, work_order_id)
    if not item or item.status != "PENDING_APPROVAL" or item.is_locked:
        raise HTTPException(status_code=409, detail="Work order is not pending completion approval")
    item.status = "APPROVAL_REJECTED"
    db.add(item)
    db.add(JobStatus(work_order_id=work_order_id, status="APPROVAL_REJECTED", timestamp=datetime.utcnow()))
    _audit(db, actor, "reject_completion", "work_order", work_order_id, {"notes": payload.notes})
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.get("/customers", response_model=list[CustomerRead])
def list_customers(
    q: str | None = None,
    limit: int = Query(default=100, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    stmt = select(Customer).where(Customer.organization_id == actor.organization_id, Customer.is_active.is_(True))
    if q:
        like = f"%{q.strip().lower()}%"
        stmt = stmt.where(func.lower(Customer.name).like(like) | func.lower(func.coalesce(Customer.account_number, "")).like(like))
    return db.scalars(stmt.order_by(Customer.name, Customer.id).limit(limit)).all()


@router.post("/equipment", response_model=EquipmentRead)
def create_equipment(payload: EquipmentCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    customer, _ = _require_tenant_service_links(db, actor, payload.customer_id, None)
    item = Equipment(organization_id=actor.organization_id, **payload.model_dump())
    if customer:
        item.customer_id = customer.id
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/equipment", response_model=list[EquipmentRead])
def list_equipment(
    customer_id: int | None = Query(default=None, ge=1),
    limit: int = Query(default=100, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    stmt = select(Equipment).where(Equipment.organization_id == actor.organization_id, Equipment.is_active.is_(True))
    if customer_id:
        stmt = stmt.where(Equipment.customer_id == customer_id)
    return db.scalars(stmt.order_by(Equipment.model, Equipment.id).limit(limit)).all()


@router.post("/work-orders/{work_order_id}/pause", response_model=WorkOrderRead)
def pause_work_order(
    work_order_id: int,
    payload: WorkOrderFlowAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    item = require_work_order_execution_scope(db, actor, work_order_id)
    if item.is_locked:
        raise HTTPException(status_code=400, detail="Completed work order cannot be paused")
    if item.status != "IN_PROGRESS":
        raise HTTPException(status_code=409, detail="Only an in-progress work order can be paused")
    item.status = "PAUSED"
    item.paused_at = datetime.utcnow()
    db.add(item)
    db.add(JobStatus(work_order_id=work_order_id, status="PAUSED", timestamp=datetime.utcnow()))
    _audit(db, actor, "pause_job", "work_order", work_order_id, {"notes": payload.notes})
    db.commit()
    db.refresh(item)
    return _work_order_read_for_actor(db, actor, item)


@router.post("/inventory/transactions", response_model=InventoryTransactionRead)
def add_inventory_transaction(
    payload: InventoryTransactionCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    effective_payload = payload.model_copy(update={"user_id": actor.user_id}) if actor.user_id else payload
    tx = create_transaction(db, effective_payload)
    _audit(
        db,
        actor,
        f"inventory_{payload.transaction_type.value}",
        "inventory_transaction",
        tx.id,
        {"part_id": payload.part_id, "qty": payload.quantity, "from": payload.from_warehouse_id, "to": payload.to_warehouse_id},
    )
    db.commit()
    db.refresh(tx)
    return tx


@router.get("/inventory/transactions", response_model=list[InventoryTransactionRead])
def list_inventory_transactions(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    return db.scalars(
        select(InventoryTransaction).order_by(InventoryTransaction.id.desc()).offset(skip).limit(limit)
    ).all()


@router.get("/inventory/balances", response_model=list[StockBalance])
def inventory_balances(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    rows = get_stock_balances(db)
    return rows[skip : skip + limit]


@router.get("/inventory/location-balances", response_model=list[LocationStockBalance])
def inventory_location_balances(
    warehouse_id: int | None = None,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    if warehouse_id is not None and not db.get(Warehouse, warehouse_id):
        raise HTTPException(status_code=404, detail="Warehouse not found")
    return get_location_stock_balances(db, warehouse_id)


@router.post("/inventory/scan", response_model=InventoryScanRead)
def scan_inventory(
    payload: InventoryScanRequest,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    if not payload.barcode and not payload.part_number:
        raise HTTPException(status_code=400, detail="barcode or part_number is required")
    part = None
    method = "barcode" if payload.barcode else "part_number"
    if payload.barcode:
        part = db.scalar(select(Part).where(Part.barcode == payload.barcode.strip()))
    if not part and payload.part_number:
        part = db.scalar(select(Part).where(Part.part_number == payload.part_number.strip()))
        if part:
            method = "part_number_fallback"
    if not part:
        return InventoryScanRead(
            matched=False, confidence=0.0, recognition_method=method,
            quantity_requested=payload.quantity, warehouse_id=payload.warehouse_id,
            location_id=payload.location_id, feedback="未匹配到物料，请拍摄清晰标签或人工选择物料。",
        )
    if payload.location_id:
        location = db.get(StorageLocation, payload.location_id)
        if not location or (payload.warehouse_id and location.warehouse_id != payload.warehouse_id):
            raise HTTPException(status_code=400, detail="Location does not belong to warehouse")
        current = get_location_stock_quantity(db, part.id, payload.location_id)
    elif payload.warehouse_id:
        current = get_stock_quantity(db, part.id, payload.warehouse_id)
    else:
        current = None
    projected = current - payload.quantity if current is not None else None
    feedback = "识别成功，库存充足。" if projected is None or projected >= 0 else f"库存不足，还差 {abs(projected)} 件。"
    _audit(db, actor, "inventory_scan", "part", part.id, {"method": method, "quantity": payload.quantity})
    db.commit()
    return InventoryScanRead(
        matched=True, confidence=1.0 if method == "barcode" else 0.95,
        recognition_method=method, part=PartRead.model_validate(part),
        quantity_requested=payload.quantity, warehouse_id=payload.warehouse_id,
        location_id=payload.location_id, current_quantity=current,
        projected_quantity=projected, feedback=feedback,
    )


@router.get("/employees/{user_id}/van-inventory", response_model=list[StockBalance])
def employee_van_inventory(
    user_id: int,
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    if actor.role == UserRole.ENGINEER and actor.user_id != user_id:
        raise HTTPException(status_code=403, detail="Can only view own van inventory")
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.ENGINEER)
    rows = get_employee_van_inventory(db, user_id)
    return rows[skip : skip + limit]


@router.get("/inventory/my-van", response_model=list[StockBalance])
def my_van_inventory(
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    if actor.role != UserRole.ENGINEER or actor.user_id is None:
        raise HTTPException(status_code=403, detail="Engineer access required")
    return get_employee_van_inventory(db, actor.user_id)[:limit]


@router.post(
    "/work-order-parts",
    response_model=WorkOrderPartRead,
    deprecated=True,
    summary="Deprecated: use /work-orders/{work_order_id}/use-part",
)
def add_work_order_part(
    payload: WorkOrderPartCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    begin_inventory_write(db)
    require_work_order_execution_scope(db, actor, payload.work_order_id)
    return use_part_for_work_order(payload.work_order_id, payload, db, actor)


@router.post(
    "/work-orders/{work_order_id}/use-part",
    response_model=WorkOrderPartRead,
    summary="Use a part on work order",
)
def use_part_for_work_order(
    work_order_id: int,
    payload: WorkOrderPartCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    begin_inventory_write(db)
    work_order = require_work_order_execution_scope(db, actor, work_order_id)
    if not work_order or work_order.is_locked or work_order.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=409, detail="Work order cannot accept parts in its current state")
    if payload.work_order_id != work_order_id:
        raise HTTPException(status_code=400, detail="Path work order ID must match payload work_order_id")
    if actor.role == UserRole.ENGINEER:
        source_warehouse = db.get(Warehouse, payload.warehouse_id)
        if (
            not source_warehouse
            or not warehouse_is_vehicle(db, source_warehouse)
            or source_warehouse.assigned_user_id != actor.user_id
        ):
            raise HTTPException(status_code=403, detail="Engineers can only use parts from their own assigned van")
    effective_payload = payload.model_copy(update={"user_id": actor.user_id}) if actor.user_id else payload
    usage = use_part_on_work_order(db, effective_payload)
    part = db.get(Part, payload.part_id)
    warehouse_quantity = get_stock_quantity(db, payload.part_id, payload.warehouse_id)
    threshold = max(part.safety_stock, part.min_stock) if part else 0
    if part and warehouse_quantity <= threshold:
        existing = db.scalar(select(InventoryNotification).where(
            InventoryNotification.part_id == part.id,
            InventoryNotification.warehouse_id == payload.warehouse_id,
            InventoryNotification.status == "open",
        ))
        if not existing:
            db.add(InventoryNotification(
                part_id=part.id, warehouse_id=payload.warehouse_id, work_order_id=work_order_id,
                message=f"{part.part_number} 使用后库存为 {warehouse_quantity}，已达到补货阈值 {threshold}。",
            ))
    _audit(db, actor, "use_part", "work_order_part", usage.id, {"work_order_id": work_order_id, "part_id": payload.part_id, "qty": payload.quantity})
    db.commit()
    return usage


@router.get("/inventory/notifications", response_model=list[InventoryNotificationRead])
def inventory_notifications(
    status: str = "open", db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    return db.scalars(select(InventoryNotification).where(InventoryNotification.status == status).order_by(InventoryNotification.id.desc()).limit(100)).all()


@router.patch("/inventory/notifications/{notification_id}", response_model=InventoryNotificationRead)
def update_inventory_notification(
    notification_id: int,
    status: str = Query(..., pattern="^(open|acknowledged|resolved)$"),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    notification = db.get(InventoryNotification, notification_id)
    if not notification:
        raise HTTPException(status_code=404, detail="Inventory notification not found")
    previous_status = notification.status
    notification.status = status
    _audit(
        db,
        actor,
        "inventory_notification_status_changed",
        "inventory_notification",
        notification.id,
        {"from_status": previous_status, "to_status": status},
    )
    db.commit()
    db.refresh(notification)
    return notification


def _replenishment_read_for_actor(
    db: Session,
    actor: Actor,
    item: ReplenishmentRequest,
) -> ReplenishmentRequestRead:
    payload = ReplenishmentRequestRead.model_validate(item).model_dump()
    part = db.get(Part, item.part_id)
    source = db.get(Warehouse, item.source_warehouse_id) if item.source_warehouse_id else None
    destination = db.get(Warehouse, item.destination_warehouse_id)
    target = db.get(User, item.target_user_id) if item.target_user_id else None
    received_device = db.get(UserDevice, item.received_device_id) if item.received_device_id else None
    work_order = db.get(WorkOrder, item.work_order_id) if item.work_order_id else None

    def user_name(user_id: int | None) -> str | None:
        user = db.get(User, user_id) if user_id else None
        return user.name if user else None

    warehouse_operator = actor.role in {UserRole.ADMIN, UserRole.WAREHOUSE}
    can_receive = False
    if item.status == "shipped":
        if item.target_user_id is not None:
            can_receive = bool(
                actor.role == UserRole.ENGINEER
                and actor.user_id == item.target_user_id
                and actor.device_verified
            )
        else:
            can_receive = warehouse_operator
    payload.update(
        part_number=part.part_number if part else None,
        part_name=part.name if part else None,
        source_warehouse_name=source.name if source else None,
        destination_warehouse_name=destination.name if destination else None,
        target_user_name=target.name if target else None,
        requested_by_name=user_name(item.requested_by),
        picking_by_name=user_name(item.picking_by),
        shipped_by_name=user_name(item.shipped_by),
        received_by_name=user_name(item.received_by),
        received_device_name=received_device.device_name if received_device else None,
        completed_by_name=user_name(item.completed_by),
        cancelled_by_name=user_name(item.cancelled_by),
        work_order_ticket_number=work_order.ticket_number if work_order else None,
        source_available_quantity=(
            get_available_stock_quantity(db, item.part_id, item.source_warehouse_id)
            if item.source_warehouse_id
            else None
        ),
        destination_quantity=(
            get_stock_quantity(db, item.part_id, item.destination_warehouse_id)
            if destination
            else 0
        ),
        can_start_picking=warehouse_operator and not item.requires_reconciliation and item.status == "requested",
        can_ship=warehouse_operator and not item.requires_reconciliation and item.status == "picking",
        can_receive=can_receive and not item.requires_reconciliation,
        can_complete=warehouse_operator and not item.requires_reconciliation and item.status == "received",
        can_cancel=warehouse_operator and not item.requires_reconciliation and item.status in {"requested", "picking"},
        can_reconcile=actor.role == UserRole.ADMIN and item.requires_reconciliation,
    )
    return ReplenishmentRequestRead(**payload)


def _validate_replenishment_destination(
    db: Session,
    destination: Warehouse,
) -> int | None:
    if not destination.is_active:
        raise HTTPException(status_code=409, detail="Destination warehouse is inactive")
    if not warehouse_is_vehicle(db, destination):
        return None
    target = db.get(User, destination.assigned_user_id) if destination.assigned_user_id else None
    if not target or not target.is_active or target.role != UserRole.ENGINEER:
        raise HTTPException(status_code=422, detail="A van destination must be assigned to an active engineer")
    return target.id


def _validate_replenishment_source(
    db: Session,
    source: Warehouse,
    destination: Warehouse,
) -> None:
    if not source.is_active or source.id == destination.id:
        raise HTTPException(status_code=400, detail="Source warehouse must be active and different from destination")
    if warehouse_is_vehicle(db, source):
        raise HTTPException(
            status_code=409,
            detail="A vehicle cannot be used as a replenishment source; use the authenticated return workflow",
        )


@router.post("/inventory/notifications/{notification_id}/create-request", response_model=ReplenishmentRequestRead)
def create_replenishment_request(
    notification_id: int,
    quantity: int = Query(default=1, ge=1),
    source_warehouse_id: int | None = Query(default=None, ge=1),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    notification = db.get(InventoryNotification, notification_id)
    if not notification:
        raise HTTPException(status_code=404, detail="Inventory notification not found")
    existing = db.scalar(
        select(ReplenishmentRequest).where(ReplenishmentRequest.notification_id == notification_id)
    )
    if existing:
        return _replenishment_read_for_actor(db, actor, existing)
    destination = db.get(Warehouse, notification.warehouse_id)
    if not destination:
        raise HTTPException(status_code=404, detail="Destination warehouse not found")
    target_user_id = _validate_replenishment_destination(db, destination)
    source = db.get(Warehouse, source_warehouse_id) if source_warehouse_id else None
    if source_warehouse_id and not source:
        raise HTTPException(status_code=404, detail="Source warehouse not found")
    if source:
        _validate_replenishment_source(db, source, destination)
    item = ReplenishmentRequest(
        notification_id=notification.id,
        part_id=notification.part_id,
        destination_warehouse_id=notification.warehouse_id,
        source_warehouse_id=source.id if source else None,
        target_user_id=target_user_id,
        quantity=quantity,
        work_order_id=notification.work_order_id,
        requested_by=actor.user_id,
    )
    notification.status = "acknowledged"
    try:
        db.add(item)
        db.flush()
        _audit(
            db,
            actor,
            "replenishment_requested",
            "replenishment_request",
            item.id,
            {
                "notification_id": notification.id,
                "part_id": item.part_id,
                "quantity": item.quantity,
                "source_warehouse_id": item.source_warehouse_id,
                "destination_warehouse_id": item.destination_warehouse_id,
                "target_user_id": item.target_user_id,
            },
        )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        existing = db.scalar(
            select(ReplenishmentRequest).where(ReplenishmentRequest.notification_id == notification_id)
        )
        if existing:
            return _replenishment_read_for_actor(db, actor, existing)
        raise HTTPException(status_code=409, detail="Replenishment request could not be created") from exc
    db.refresh(item)
    return _replenishment_read_for_actor(db, actor, item)


@router.post("/inventory/replenishment-requests", response_model=ReplenishmentRequestRead)
def create_manual_replenishment_request(
    payload: ReplenishmentRequestCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    client_request_id = payload.client_request_id.strip()
    request_reason = payload.reason.strip()
    if len(request_reason) < 3:
        raise HTTPException(status_code=422, detail="Request reason must be at least 3 non-whitespace characters")

    def matches_existing(candidate: ReplenishmentRequest) -> bool:
        return (
            candidate.part_id == payload.part_id
            and candidate.destination_warehouse_id == payload.destination_warehouse_id
            and candidate.source_warehouse_id == payload.source_warehouse_id
            and candidate.quantity == payload.quantity
            and (candidate.request_reason or "") == request_reason
        )

    existing = db.scalar(
        select(ReplenishmentRequest).where(
            ReplenishmentRequest.client_request_id == client_request_id,
        )
    )
    if existing:
        if not matches_existing(existing):
            raise HTTPException(status_code=409, detail="client_request_id was already used for another request")
        return _replenishment_read_for_actor(db, actor, existing)

    part = db.get(Part, payload.part_id)
    destination = db.get(Warehouse, payload.destination_warehouse_id)
    source = db.get(Warehouse, payload.source_warehouse_id) if payload.source_warehouse_id else None
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    if not destination:
        raise HTTPException(status_code=404, detail="Destination warehouse not found")
    target_user_id = _validate_replenishment_destination(db, destination)
    if target_user_id is None:
        raise HTTPException(status_code=422, detail="Manual replenishment destination must be an assigned vehicle")
    if payload.source_warehouse_id and not source:
        raise HTTPException(status_code=404, detail="Source warehouse not found")
    if source:
        _validate_replenishment_source(db, source, destination)

    item = ReplenishmentRequest(
        client_request_id=client_request_id,
        request_reason=request_reason,
        part_id=part.id,
        destination_warehouse_id=destination.id,
        source_warehouse_id=source.id if source else None,
        target_user_id=target_user_id,
        quantity=payload.quantity,
        requested_by=actor.user_id,
    )
    try:
        db.add(item)
        db.flush()
        _audit(
            db,
            actor,
            "replenishment_requested",
            "replenishment_request",
            item.id,
            {
                "origin": "manual",
                "client_request_id": client_request_id,
                "reason": item.request_reason,
                "part_id": item.part_id,
                "quantity": item.quantity,
                "source_warehouse_id": item.source_warehouse_id,
                "destination_warehouse_id": item.destination_warehouse_id,
                "target_user_id": item.target_user_id,
            },
        )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        existing = db.scalar(
            select(ReplenishmentRequest).where(
                ReplenishmentRequest.client_request_id == client_request_id,
            )
        )
        if existing:
            if matches_existing(existing):
                return _replenishment_read_for_actor(db, actor, existing)
            raise HTTPException(status_code=409, detail="client_request_id was already used for another request")
        raise HTTPException(status_code=409, detail="Replenishment request could not be created") from exc
    db.refresh(item)
    return _replenishment_read_for_actor(db, actor, item)


@router.get("/inventory/replenishment-requests", response_model=list[ReplenishmentRequestRead])
def list_replenishment_requests(
    status: str | None = Query(default=None, pattern="^(requested|picking|shipped|received|completed|cancelled)$"),
    limit: int = Query(default=100, ge=1, le=200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    stmt = select(ReplenishmentRequest).order_by(ReplenishmentRequest.id.desc())
    if actor.role == UserRole.ENGINEER:
        stmt = stmt.where(ReplenishmentRequest.target_user_id == actor.user_id)
    if status:
        stmt = stmt.where(ReplenishmentRequest.status == status)
    rows = db.scalars(stmt.limit(limit)).all()
    return [_replenishment_read_for_actor(db, actor, item) for item in rows]


@router.post(
    "/inventory/replenishment-requests/{request_id}/reconcile",
    response_model=ReplenishmentRequestRead,
)
def reconcile_replenishment_request(
    request_id: int,
    payload: ReplenishmentRequestReconcile,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    if actor.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Administrator access required")
    begin_inventory_write(db)
    item = db.scalar(
        select(ReplenishmentRequest)
        .where(
            ReplenishmentRequest.id == request_id,
            ReplenishmentRequest.organization_id == actor.organization_id,
        )
        .with_for_update()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Replenishment request not found")
    if not item.requires_reconciliation:
        return _replenishment_read_for_actor(db, actor, item)
    if item.version != payload.expected_version:
        raise HTTPException(status_code=409, detail="Replenishment request changed; refresh before reconciling")
    reason = payload.reason.strip()
    if len(reason) < 3:
        raise HTTPException(status_code=422, detail="Reconciliation reason must be at least 3 characters")
    _require_account_reauthentication(db, actor, payload.account_password)
    linked_movements = db.scalars(
        select(InventoryTransaction).where(InventoryTransaction.replenishment_request_id == item.id)
    ).all()
    if linked_movements:
        raise HTTPException(
            status_code=409,
            detail="Linked inventory movements require a dedicated stock correction, not historical reconciliation",
        )
    if payload.resolution == "reset_requested":
        if item.status != "requested":
            raise HTTPException(status_code=409, detail="Only a reopened requested record can use reset_requested")
        if item.notification_id:
            notification = db.get(InventoryNotification, item.notification_id)
            if notification:
                notification.status = "open"
    elif item.status != "completed":
        raise HTTPException(status_code=409, detail="accept_historical is only valid for a legacy completed record")
    elif item.notification_id:
        notification = db.get(InventoryNotification, item.notification_id)
        if notification:
            notification.status = "resolved"

    previous_version = item.version
    item.requires_reconciliation = False
    item.version += 1
    _audit(
        db,
        actor,
        "replenishment_reconciled",
        "replenishment_request",
        item.id,
        {
            "resolution": payload.resolution,
            "reason": reason,
            "previous_version": previous_version,
            "new_version": item.version,
            "status": item.status,
        },
    )
    db.commit()
    db.refresh(item)
    return _replenishment_read_for_actor(db, actor, item)


@router.post("/inventory/replenishment-requests/{request_id}/actions", response_model=ReplenishmentRequestRead)
def act_on_replenishment_request(
    request_id: int,
    payload: ReplenishmentRequestAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    begin_inventory_write(db)
    item = db.scalar(
        select(ReplenishmentRequest)
        .where(
            ReplenishmentRequest.id == request_id,
            ReplenishmentRequest.organization_id == actor.organization_id,
        )
        .with_for_update()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Replenishment request not found")
    if item.requires_reconciliation:
        raise HTTPException(
            status_code=409,
            detail="Historical replenishment requires administrator reconciliation before workflow actions",
        )
    target_status = {
        "start_picking": "picking",
        "ship": "shipped",
        "receive": "received",
        "complete": "completed",
        "cancel": "cancelled",
    }[payload.action]
    warehouse_operator = actor.role in {UserRole.ADMIN, UserRole.WAREHOUSE}
    if payload.action in {"start_picking", "ship", "complete", "cancel"} and not warehouse_operator:
        raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
    if payload.action == "receive":
        if item.target_user_id is not None:
            if actor.role != UserRole.ENGINEER or actor.user_id != item.target_user_id:
                raise HTTPException(status_code=403, detail="Only the destination engineer can receive this shipment")
            require_bound_device(actor)
            if item.status == "received" and item.received_device_id != actor.device_record_id:
                raise HTTPException(status_code=403, detail="Shipment was received on another registered device")
        elif not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
    if item.status == target_status:
        return _replenishment_read_for_actor(db, actor, item)
    if item.version != payload.expected_version:
        raise HTTPException(status_code=409, detail="Replenishment request changed; refresh before continuing")

    previous_status = item.status
    transaction_id: int | None = None
    now = datetime.utcnow()
    if payload.action == "start_picking":
        if not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
        if item.status != "requested":
            raise HTTPException(status_code=409, detail="Only requested replenishments can start picking")
        if (
            item.source_warehouse_id is not None
            and payload.source_warehouse_id is not None
            and payload.source_warehouse_id != item.source_warehouse_id
        ):
            raise HTTPException(
                status_code=409,
                detail="The assigned source warehouse cannot be replaced during picking; cancel and recreate the request",
            )
        source_id = item.source_warehouse_id or payload.source_warehouse_id
        source = db.get(Warehouse, source_id) if source_id else None
        destination = db.get(Warehouse, item.destination_warehouse_id)
        if not source:
            raise HTTPException(status_code=422, detail="Select a source warehouse before picking")
        if not destination:
            raise HTTPException(status_code=404, detail="Destination warehouse not found")
        _validate_replenishment_source(db, source, destination)
        db.scalar(select(Part).where(Part.id == item.part_id).with_for_update())
        if get_available_stock_quantity(db, item.part_id, source.id) < item.quantity:
            raise HTTPException(status_code=409, detail="Insufficient unreserved source stock")
        current_target = _validate_replenishment_destination(db, destination)
        if item.target_user_id not in {None, current_target}:
            raise HTTPException(status_code=409, detail="Destination van ownership changed; cancel and recreate the request")
        item.source_warehouse_id = source.id
        item.target_user_id = current_target
        item.picking_by = actor.user_id
        item.picking_at = now

    elif payload.action == "ship":
        if not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
        if item.status != "picking" or not item.source_warehouse_id:
            raise HTTPException(status_code=409, detail="Only a picking request with a source can be shipped")
        source = db.get(Warehouse, item.source_warehouse_id)
        destination = db.get(Warehouse, item.destination_warehouse_id)
        if not source or not source.is_active:
            raise HTTPException(status_code=409, detail="Source warehouse is no longer active")
        if not destination:
            raise HTTPException(status_code=404, detail="Destination warehouse not found")
        _validate_replenishment_source(db, source, destination)
        current_target = _validate_replenishment_destination(db, destination)
        if current_target != item.target_user_id:
            raise HTTPException(status_code=409, detail="Destination custody changed; cancel and recreate the request")
        part = db.scalar(select(Part).where(Part.id == item.part_id).with_for_update())
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        if get_stock_quantity(db, item.part_id, item.source_warehouse_id) < item.quantity:
            raise HTTPException(status_code=409, detail="Source stock is no longer sufficient")
        transaction = InventoryTransaction(
            part_id=item.part_id,
            transaction_type=TransactionType.OUTBOUND,
            quantity=item.quantity,
            from_warehouse_id=item.source_warehouse_id,
            work_order_id=item.work_order_id,
            replenishment_request_id=item.id,
            movement_stage="ship",
            user_id=actor.user_id,
            unit_cost=part.default_cost,
            notes=f"Replenishment #{item.id} shipped",
        )
        db.add(transaction)
        db.flush()
        transaction_id = transaction.id
        item.shipment_transaction_id = transaction.id
        item.shipped_by = actor.user_id
        item.shipped_at = now

    elif payload.action == "receive":
        if item.status != "shipped":
            raise HTTPException(status_code=409, detail="Only shipped replenishments can be received")
        destination = db.get(Warehouse, item.destination_warehouse_id)
        if not destination:
            raise HTTPException(status_code=404, detail="Destination warehouse not found")
        current_target = _validate_replenishment_destination(db, destination)
        if current_target != item.target_user_id:
            raise HTTPException(status_code=409, detail="Destination custody changed after shipment; warehouse reconciliation is required")
        if item.target_user_id is not None:
            if actor.role != UserRole.ENGINEER or actor.user_id != item.target_user_id:
                raise HTTPException(status_code=403, detail="Only the destination engineer can receive this shipment")
            require_bound_device(actor)
            if destination.assigned_user_id != actor.user_id:
                raise HTTPException(status_code=409, detail="Destination van is no longer assigned to this engineer")
            _require_account_reauthentication(db, actor, payload.account_password)
        elif not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
        shipment_transaction = (
            db.get(InventoryTransaction, item.shipment_transaction_id)
            if item.shipment_transaction_id
            else None
        )
        if (
            not shipment_transaction
            or shipment_transaction.replenishment_request_id != item.id
            or shipment_transaction.movement_stage != "ship"
            or shipment_transaction.transaction_type != TransactionType.OUTBOUND
            or shipment_transaction.part_id != item.part_id
            or shipment_transaction.quantity != item.quantity
            or shipment_transaction.from_warehouse_id != item.source_warehouse_id
        ):
            raise HTTPException(status_code=409, detail="Shipment ledger is incomplete; warehouse reconciliation is required")
        part = db.scalar(select(Part).where(Part.id == item.part_id).with_for_update())
        if not part:
            raise HTTPException(status_code=404, detail="Part not found")
        transaction = InventoryTransaction(
            part_id=item.part_id,
            transaction_type=TransactionType.INBOUND,
            quantity=item.quantity,
            to_warehouse_id=item.destination_warehouse_id,
            work_order_id=item.work_order_id,
            replenishment_request_id=item.id,
            movement_stage="receive",
            user_id=actor.user_id,
            unit_cost=shipment_transaction.unit_cost,
            notes=f"Replenishment #{item.id} received",
        )
        db.add(transaction)
        db.flush()
        transaction_id = transaction.id
        item.receipt_transaction_id = transaction.id
        item.received_by = actor.user_id
        item.received_device_id = actor.device_record_id
        item.received_at = now

    elif payload.action == "complete":
        if not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
        if item.status != "received" or item.receipt_transaction_id is None:
            raise HTTPException(status_code=409, detail="Only a received shipment can be completed")
        receipt_transaction = db.get(InventoryTransaction, item.receipt_transaction_id)
        if (
            not receipt_transaction
            or receipt_transaction.replenishment_request_id != item.id
            or receipt_transaction.movement_stage != "receive"
            or receipt_transaction.transaction_type != TransactionType.INBOUND
            or receipt_transaction.part_id != item.part_id
            or receipt_transaction.quantity != item.quantity
            or receipt_transaction.to_warehouse_id != item.destination_warehouse_id
        ):
            raise HTTPException(status_code=409, detail="Receipt ledger is incomplete; warehouse reconciliation is required")
        item.completed_by = actor.user_id
        item.completed_at = now
        if item.notification_id:
            notification = db.get(InventoryNotification, item.notification_id)
            if notification:
                notification.status = "resolved"

    else:
        if not warehouse_operator:
            raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
        if item.status not in {"requested", "picking"}:
            raise HTTPException(status_code=409, detail="Only requested or picking replenishments can be cancelled")
        if not payload.reason or len(payload.reason.strip()) < 3:
            raise HTTPException(status_code=422, detail="Cancellation reason must be at least 3 characters")
        item.cancelled_by = actor.user_id
        item.cancelled_at = now
        item.cancellation_reason = payload.reason.strip()
        if item.notification_id:
            notification = db.get(InventoryNotification, item.notification_id)
            if notification:
                notification.status = "resolved"

    item.status = target_status
    item.version += 1
    db.add(item)
    _audit(
        db,
        actor,
        f"replenishment_{payload.action}",
        "replenishment_request",
        item.id,
        {
            "from_status": previous_status,
            "to_status": target_status,
            "previous_version": payload.expected_version,
            "new_version": item.version,
            "part_id": item.part_id,
            "quantity": item.quantity,
            "source_warehouse_id": item.source_warehouse_id,
            "destination_warehouse_id": item.destination_warehouse_id,
            "target_user_id": item.target_user_id,
            "inventory_transaction_id": transaction_id,
        },
    )
    db.commit()
    db.refresh(item)
    return _replenishment_read_for_actor(db, actor, item)


@router.patch(
    "/inventory/replenishment-requests/{request_id}",
    response_model=ReplenishmentRequestRead,
    deprecated=True,
)
def update_replenishment_request(request_id: int):
    raise HTTPException(status_code=410, detail="Use the authenticated replenishment action endpoint")


def _validate_vehicle_return_warehouses(
    db: Session,
    source: Warehouse,
    destination: Warehouse,
    engineer_id: int,
) -> None:
    if not source.is_active or not warehouse_is_vehicle(db, source):
        raise HTTPException(status_code=409, detail="Return source must be an active engineer vehicle")
    if source.assigned_user_id != engineer_id:
        raise HTTPException(status_code=403, detail="Return source is not assigned to this engineer")
    if not destination.is_active or warehouse_is_vehicle(db, destination):
        raise HTTPException(status_code=409, detail="Return destination must be an active non-vehicle warehouse")
    if source.id == destination.id:
        raise HTTPException(status_code=400, detail="Return source and destination must be different")


def _vehicle_return_read_for_actor(
    db: Session,
    actor: Actor,
    item: VehicleReturnRequest,
) -> VehicleReturnRequestRead:
    payload = VehicleReturnRequestRead.model_validate(item).model_dump()
    part = db.get(Part, item.part_id)
    source = db.get(Warehouse, item.source_warehouse_id)
    destination = db.get(Warehouse, item.destination_warehouse_id)
    requested_device = db.get(UserDevice, item.requested_device_id)
    shipped_device = db.get(UserDevice, item.shipped_device_id) if item.shipped_device_id else None

    def user_name(user_id: int | None) -> str | None:
        user = db.get(User, user_id) if user_id else None
        return user.name if user else None

    warehouse_operator = actor.role in {UserRole.ADMIN, UserRole.WAREHOUSE}
    is_engineer_owner = bool(actor.role == UserRole.ENGINEER and actor.user_id == item.engineer_id)
    payload.update(
        part_number=part.part_number if part else None,
        part_name=part.name if part else None,
        source_warehouse_name=source.name if source else None,
        destination_warehouse_name=destination.name if destination else None,
        engineer_name=user_name(item.engineer_id),
        requested_by_name=user_name(item.requested_by),
        requested_device_name=requested_device.device_name if requested_device else None,
        approved_by_name=user_name(item.approved_by),
        shipped_by_name=user_name(item.shipped_by),
        shipped_device_name=shipped_device.device_name if shipped_device else None,
        received_by_name=user_name(item.received_by),
        cancelled_by_name=user_name(item.cancelled_by),
        source_quantity=get_stock_quantity(db, item.part_id, item.source_warehouse_id),
        destination_quantity=get_stock_quantity(db, item.part_id, item.destination_warehouse_id),
        can_approve=warehouse_operator and item.status == "requested",
        can_ship=is_engineer_owner and actor.device_verified and item.status == "approved",
        can_receive=warehouse_operator and item.status == "shipped",
        can_cancel=(warehouse_operator or is_engineer_owner) and item.status in {"requested", "approved"},
    )
    return VehicleReturnRequestRead(**payload)


@router.get("/inventory/vehicle-return-destinations", response_model=list[WarehouseRead])
def vehicle_return_destinations(
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    rows = db.scalars(select(Warehouse).where(Warehouse.is_active.is_(True)).order_by(Warehouse.name)).all()
    return [warehouse for warehouse in rows if not warehouse_is_vehicle(db, warehouse)]


@router.post("/inventory/vehicle-returns", response_model=VehicleReturnRequestRead)
def create_vehicle_return_request(
    payload: VehicleReturnRequestCreate,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    if actor.role != UserRole.ENGINEER or actor.user_id is None:
        raise HTTPException(status_code=403, detail="Only an engineer can request a vehicle return")
    require_bound_device(actor)
    client_request_id = payload.client_request_id.strip()
    reason = payload.reason.strip()
    if len(reason) < 3:
        raise HTTPException(status_code=422, detail="Return reason must be at least 3 non-whitespace characters")

    def matches_existing(candidate: VehicleReturnRequest) -> bool:
        return (
            candidate.part_id == payload.part_id
            and candidate.source_warehouse_id == payload.source_warehouse_id
            and candidate.destination_warehouse_id == payload.destination_warehouse_id
            and candidate.quantity == payload.quantity
            and candidate.reason == reason
            and candidate.engineer_id == actor.user_id
        )

    existing = db.scalar(
        select(VehicleReturnRequest).where(VehicleReturnRequest.client_request_id == client_request_id)
    )
    if existing:
        if not matches_existing(existing):
            raise HTTPException(status_code=409, detail="client_request_id was already used for another return")
        return _vehicle_return_read_for_actor(db, actor, existing)

    part = db.get(Part, payload.part_id)
    source = db.get(Warehouse, payload.source_warehouse_id)
    destination = db.get(Warehouse, payload.destination_warehouse_id)
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")
    if not source or not destination:
        raise HTTPException(status_code=404, detail="Return warehouse not found")
    _validate_vehicle_return_warehouses(db, source, destination, actor.user_id)
    if get_stock_quantity(db, part.id, source.id) < payload.quantity:
        raise HTTPException(status_code=409, detail="Vehicle stock is insufficient for this return")

    item = VehicleReturnRequest(
        client_request_id=client_request_id,
        part_id=part.id,
        source_warehouse_id=source.id,
        destination_warehouse_id=destination.id,
        engineer_id=actor.user_id,
        quantity=payload.quantity,
        reason=reason,
        requested_by=actor.user_id,
        requested_device_id=actor.device_record_id,
    )
    try:
        db.add(item)
        db.flush()
        _audit(
            db,
            actor,
            "vehicle_return_requested",
            "vehicle_return_request",
            item.id,
            {
                "part_id": item.part_id,
                "quantity": item.quantity,
                "source_warehouse_id": item.source_warehouse_id,
                "destination_warehouse_id": item.destination_warehouse_id,
                "engineer_id": item.engineer_id,
                "client_request_id": item.client_request_id,
                "reason": item.reason,
            },
        )
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        existing = db.scalar(
            select(VehicleReturnRequest).where(VehicleReturnRequest.client_request_id == client_request_id)
        )
        if existing and matches_existing(existing):
            return _vehicle_return_read_for_actor(db, actor, existing)
        if existing:
            raise HTTPException(status_code=409, detail="client_request_id was already used for another return")
        raise HTTPException(status_code=409, detail="Vehicle return request could not be created") from exc
    db.refresh(item)
    return _vehicle_return_read_for_actor(db, actor, item)


@router.get("/inventory/vehicle-returns", response_model=list[VehicleReturnRequestRead])
def list_vehicle_return_requests(
    status: str | None = Query(default=None, pattern="^(requested|approved|shipped|received|cancelled)$"),
    limit: int = Query(default=100, ge=1, le=200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE, UserRole.ENGINEER)
    stmt = select(VehicleReturnRequest).order_by(VehicleReturnRequest.id.desc())
    if actor.role == UserRole.ENGINEER:
        stmt = stmt.where(VehicleReturnRequest.engineer_id == actor.user_id)
    if status:
        stmt = stmt.where(VehicleReturnRequest.status == status)
    rows = db.scalars(stmt.limit(limit)).all()
    return [_vehicle_return_read_for_actor(db, actor, item) for item in rows]


@router.post("/inventory/vehicle-returns/{request_id}/actions", response_model=VehicleReturnRequestRead)
def act_on_vehicle_return_request(
    request_id: int,
    payload: VehicleReturnRequestAction,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    begin_inventory_write(db)
    item = db.scalar(
        select(VehicleReturnRequest)
        .where(
            VehicleReturnRequest.id == request_id,
            VehicleReturnRequest.organization_id == actor.organization_id,
        )
        .with_for_update()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Vehicle return request not found")

    warehouse_operator = actor.role in {UserRole.ADMIN, UserRole.WAREHOUSE}
    engineer_owner = bool(actor.role == UserRole.ENGINEER and actor.user_id == item.engineer_id)
    target_status = {
        "approve": "approved",
        "ship": "shipped",
        "receive": "received",
        "cancel": "cancelled",
    }[payload.action]
    if payload.action in {"approve", "receive"} and not warehouse_operator:
        raise HTTPException(status_code=403, detail="Warehouse or administrator access required")
    if payload.action == "ship":
        if not engineer_owner:
            raise HTTPException(status_code=403, detail="Only the vehicle owner can hand over this return")
        require_bound_device(actor)
        if item.status == "shipped" and item.shipped_device_id != actor.device_record_id:
            raise HTTPException(status_code=403, detail="Return was handed over on another registered device")
    if payload.action == "cancel" and not (warehouse_operator or engineer_owner):
        raise HTTPException(status_code=403, detail="Only the vehicle owner or warehouse can cancel this return")
    if item.status == target_status:
        return _vehicle_return_read_for_actor(db, actor, item)
    if item.version != payload.expected_version:
        raise HTTPException(status_code=409, detail="Vehicle return changed; refresh before continuing")

    part = db.scalar(select(Part).where(Part.id == item.part_id).with_for_update())
    source = db.get(Warehouse, item.source_warehouse_id)
    destination = db.get(Warehouse, item.destination_warehouse_id)
    if not part or not source or not destination:
        raise HTTPException(status_code=409, detail="Return inventory references are incomplete")

    previous_status = item.status
    transaction_id: int | None = None
    now = datetime.utcnow()
    if payload.action == "approve":
        if item.status != "requested":
            raise HTTPException(status_code=409, detail="Only a requested return can be approved")
        _validate_vehicle_return_warehouses(db, source, destination, item.engineer_id)
        if get_available_stock_quantity(db, item.part_id, item.source_warehouse_id) < item.quantity:
            raise HTTPException(status_code=409, detail="Vehicle stock is no longer available for approval")
        item.approved_by = actor.user_id
        item.approved_at = now

    elif payload.action == "ship":
        if item.status != "approved":
            raise HTTPException(status_code=409, detail="Only an approved return can be handed over")
        _validate_vehicle_return_warehouses(db, source, destination, item.engineer_id)
        if source.assigned_user_id != actor.user_id:
            raise HTTPException(status_code=409, detail="Vehicle assignment changed after approval")
        _require_account_reauthentication(db, actor, payload.account_password)
        if get_stock_quantity(db, item.part_id, item.source_warehouse_id) < item.quantity:
            raise HTTPException(status_code=409, detail="Vehicle stock is insufficient for handover")
        transaction = InventoryTransaction(
            part_id=item.part_id,
            transaction_type=TransactionType.OUTBOUND,
            quantity=item.quantity,
            from_warehouse_id=item.source_warehouse_id,
            vehicle_return_request_id=item.id,
            movement_stage="return_ship",
            user_id=actor.user_id,
            unit_cost=part.default_cost,
            notes=f"Vehicle return #{item.id} handed over",
        )
        db.add(transaction)
        db.flush()
        transaction_id = transaction.id
        item.shipment_transaction_id = transaction.id
        item.shipped_by = actor.user_id
        item.shipped_device_id = actor.device_record_id
        item.shipped_at = now

    elif payload.action == "receive":
        if item.status != "shipped":
            raise HTTPException(status_code=409, detail="Only a handed-over return can be received")
        if not destination.is_active or warehouse_is_vehicle(db, destination):
            raise HTTPException(status_code=409, detail="Return destination is no longer an active warehouse")
        shipment = db.get(InventoryTransaction, item.shipment_transaction_id) if item.shipment_transaction_id else None
        if (
            not shipment
            or shipment.vehicle_return_request_id != item.id
            or shipment.movement_stage != "return_ship"
            or shipment.transaction_type != TransactionType.OUTBOUND
            or shipment.part_id != item.part_id
            or shipment.quantity != item.quantity
            or shipment.from_warehouse_id != item.source_warehouse_id
        ):
            raise HTTPException(status_code=409, detail="Return shipment ledger is incomplete")
        transaction = InventoryTransaction(
            part_id=item.part_id,
            transaction_type=TransactionType.INBOUND,
            quantity=item.quantity,
            to_warehouse_id=item.destination_warehouse_id,
            vehicle_return_request_id=item.id,
            movement_stage="return_receive",
            user_id=actor.user_id,
            unit_cost=shipment.unit_cost,
            notes=f"Vehicle return #{item.id} received",
        )
        db.add(transaction)
        db.flush()
        transaction_id = transaction.id
        item.receipt_transaction_id = transaction.id
        item.received_by = actor.user_id
        item.received_at = now

    else:
        if item.status not in {"requested", "approved"}:
            raise HTTPException(status_code=409, detail="Only requested or approved returns can be cancelled")
        if not payload.reason or len(payload.reason.strip()) < 3:
            raise HTTPException(status_code=422, detail="Cancellation reason must be at least 3 characters")
        item.cancelled_by = actor.user_id
        item.cancelled_at = now
        item.cancellation_reason = payload.reason.strip()

    item.status = target_status
    item.version += 1
    _audit(
        db,
        actor,
        f"vehicle_return_{payload.action}",
        "vehicle_return_request",
        item.id,
        {
            "from_status": previous_status,
            "to_status": target_status,
            "previous_version": payload.expected_version,
            "new_version": item.version,
            "part_id": item.part_id,
            "quantity": item.quantity,
            "source_warehouse_id": item.source_warehouse_id,
            "destination_warehouse_id": item.destination_warehouse_id,
            "engineer_id": item.engineer_id,
            "inventory_transaction_id": transaction_id,
            "reason": item.cancellation_reason if payload.action == "cancel" else None,
        },
    )
    db.commit()
    db.refresh(item)
    return _vehicle_return_read_for_actor(db, actor, item)


@router.get("/work-orders/{work_order_id}/part-recommendations", response_model=list[WorkOrderPartRecommendation])
def work_order_part_recommendations(
    work_order_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_work_order_scope(db, actor, work_order_id)
    work_order = db.get(WorkOrder, work_order_id)
    if not work_order:
        raise HTTPException(status_code=404, detail="Work order not found")
    memories = db.scalars(select(WorkOrderPartMemory).where(
        or_(
            and_(WorkOrderPartMemory.machine_type == work_order.machine_type, WorkOrderPartMemory.job_type == work_order.job_type),
            WorkOrderPartMemory.machine_type == work_order.machine_type,
            WorkOrderPartMemory.job_type == work_order.job_type,
        )
    ).order_by(WorkOrderPartMemory.usage_count.desc(), WorkOrderPartMemory.total_quantity.desc()).limit(20)).all()
    recommendations = []
    for memory in memories:
        part = db.get(Part, memory.part_id)
        if part:
            average = max(1, round(memory.total_quantity / memory.usage_count))
            exact = memory.machine_type == work_order.machine_type and memory.job_type == work_order.job_type
            basis = "相同机型和工单类型" if exact else ("相同机型" if memory.machine_type == work_order.machine_type else "相同工单类型")
            recommendations.append(WorkOrderPartRecommendation(
                part=PartRead.model_validate(part), recommended_quantity=average,
                usage_count=memory.usage_count, total_quantity=memory.total_quantity,
                reason=f"基于{basis}：历史上 {memory.usage_count} 个类似工单使用过，平均每单 {average} 件。",
            ))
    return recommendations


@router.get("/work-order-parts", response_model=list[WorkOrderPartRead])
def list_work_order_parts(
    work_order_id: int | None = Query(default=None, ge=1),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    stmt = select(WorkOrderPart).order_by(WorkOrderPart.id.desc())
    if work_order_id is not None:
        require_work_order_scope(db, actor, work_order_id)
        stmt = stmt.where(WorkOrderPart.work_order_id == work_order_id)
    elif actor.role == UserRole.ENGINEER and actor.user_id:
        stmt = stmt.where(WorkOrderPart.user_id == actor.user_id)
    elif actor.role not in {UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE}:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    return db.scalars(stmt.offset(skip).limit(limit)).all()


@router.post("/qc-pictures", response_model=QCPictureRead)
def create_qc_picture(
    payload: QCPictureCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    work_order = require_work_order_execution_scope(db, actor, payload.work_order_id)
    if not work_order:
        raise HTTPException(status_code=404, detail="Work order not found")
    if work_order.is_locked or work_order.status == "PENDING_APPROVAL":
        raise HTTPException(status_code=400, detail="Work order cannot accept more photos in its current state")
    payload.uploaded_by = actor.user_id
    item = QCPicture(**payload.model_dump())
    db.add(item)
    _audit(db, actor, "upload_qc_picture", "qc_picture", None, {"work_order_id": payload.work_order_id})
    db.commit()
    db.refresh(item)
    return item


@router.get("/qc-pictures", response_model=list[QCPictureRead])
def list_qc_pictures(
    work_order_id: int | None = Query(default=None, ge=1),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    stmt = select(QCPicture).order_by(QCPicture.id.desc())
    if work_order_id:
        require_work_order_scope(db, actor, work_order_id)
        stmt = stmt.where(QCPicture.work_order_id == work_order_id)
    elif actor.role == UserRole.ENGINEER:
        raise HTTPException(status_code=403, detail="work_order_id is required for technician scope")
    return db.scalars(stmt.offset(skip).limit(limit)).all()


@router.post("/job-status", response_model=JobStatusRead)
def create_job_status(
    payload: JobStatusCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    work_order = require_work_order_execution_scope(db, actor, payload.work_order_id)
    if work_order and (work_order.is_locked or work_order.status == "PENDING_APPROVAL"):
        raise HTTPException(status_code=400, detail="Work order is locked and cannot be edited")
    if payload.status.strip().upper() in {"COMPLETED", "PENDING_APPROVAL", "APPROVAL_REJECTED"}:
        raise HTTPException(status_code=400, detail="Reserved status requires the completion workflow")
    data = payload.model_dump()
    if not data.get("timestamp"):
        data["timestamp"] = datetime.utcnow()
    item = JobStatus(**data)
    db.add(item)
    _audit(db, actor, "update_status", "job_status", None, {"work_order_id": payload.work_order_id, "status": data.get("status")})
    db.commit()
    db.refresh(item)
    return item


@router.get("/job-status", response_model=list[JobStatusRead])
def list_job_status(
    work_order_id: int | None = Query(default=None, ge=1),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    stmt = select(JobStatus).order_by(JobStatus.id.desc())
    if work_order_id:
        require_work_order_scope(db, actor, work_order_id)
        stmt = stmt.where(JobStatus.work_order_id == work_order_id)
    elif actor.role == UserRole.ENGINEER:
        raise HTTPException(status_code=403, detail="work_order_id is required for technician scope")
    return db.scalars(stmt.offset(skip).limit(limit)).all()


@router.post("/return-equipments", response_model=ReturnEquipmentRead)
def create_return_equipment(
    payload: ReturnEquipmentCreate, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    work_order = require_work_order_execution_scope(db, actor, payload.work_order_id)
    if work_order and (work_order.is_locked or work_order.status == "PENDING_APPROVAL"):
        raise HTTPException(status_code=400, detail="Work order is locked and cannot be edited")
    item = ReturnEquipment(**payload.model_dump())
    db.add(item)
    _audit(
        db,
        actor,
        "return_equipment",
        "return_equipment",
        None,
        {"work_order_id": payload.work_order_id, "equipment_type": payload.equipment_type, "quantity": payload.quantity},
    )
    db.commit()
    db.refresh(item)
    return item


@router.get("/audit-logs")
def list_audit_logs(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    rows = db.scalars(select(AuditLog).order_by(AuditLog.id.desc()).offset(skip).limit(limit)).all()
    return [
        {
            "id": row.id,
            "user_id": row.user_id,
            "action": row.action,
            "entity_type": row.entity_type,
            "entity_id": row.entity_id,
            "timestamp": row.timestamp.isoformat() if row.timestamp else None,
            "metadata": row.metadata_json,
        }
        for row in rows
    ]


@router.get("/return-equipments", response_model=list[ReturnEquipmentRead])
def list_return_equipments(
    work_order_id: int | None = Query(default=None, ge=1),
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    stmt = select(ReturnEquipment).order_by(ReturnEquipment.id.desc())
    if work_order_id:
        require_work_order_scope(db, actor, work_order_id)
        stmt = stmt.where(ReturnEquipment.work_order_id == work_order_id)
    elif actor.role == UserRole.ENGINEER:
        raise HTTPException(status_code=403, detail="work_order_id is required for technician scope")
    return db.scalars(stmt.offset(skip).limit(limit)).all()


@router.get("/work-orders/{work_order_id}/profit", response_model=WorkOrderProfit)
def work_order_profit(work_order_id: int, db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    work_order = db.get(WorkOrder, work_order_id)
    if not work_order:
        raise HTTPException(status_code=404, detail="Work order not found")

    parts_cost = get_work_order_parts_cost(db, work_order_id)
    return WorkOrderProfit(
        work_order_id=work_order_id,
        ticket_number=work_order.ticket_number,
        wo_number=work_order.wo_number,
        revenue=work_order.revenue,
        labor_cost=work_order.labor_cost,
        parts_cost=parts_cost,
        profit=work_order.revenue - work_order.labor_cost - parts_cost,
    )


@router.get("/export/inventory.xlsx")
def export_inventory_excel(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    balances = get_stock_balances(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Balance"
    ws.append(["Part ID", "Part Number", "Part Name", "Warehouse ID", "Warehouse", "Quantity", "Safety Stock", "Low Stock"])
    for row in balances:
        ws.append([
            row.part_id,
            row.part_number,
            row.part_name,
            row.warehouse_id,
            row.warehouse_name,
            row.quantity,
            row.safety_stock,
            "YES" if row.is_low_stock else "NO",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory.xlsx"},
    )


@router.get("/export/parts.xlsx")
def export_parts_excel(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    rows = db.scalars(select(Part).order_by(Part.id.asc())).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"
    ws.append(["part_number", "name", "unit", "default_cost", "safety_stock", "supplier", "notes"])
    for row in rows:
        ws.append([row.part_number, row.name, row.unit, row.default_cost, row.safety_stock, row.supplier, row.notes])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=parts.xlsx"},
    )


@router.post("/imports/parts/preview", response_model=ImportBatchRead)
async def preview_parts_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    content = await file.read(settings.max_import_upload_bytes + 1)
    if len(content) > settings.max_import_upload_bytes:
        raise HTTPException(status_code=413, detail="Import file exceeds the configured upload limit")

    file_hash = sha256(content).hexdigest()
    existing_batch = db.scalar(
        select(ImportBatch)
        .where(ImportBatch.import_type == "parts", ImportBatch.file_sha256 == file_hash)
        .order_by(ImportBatch.id.desc())
    )
    if existing_batch:
        return _import_batch_read(existing_batch)

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        worksheet = workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The uploaded workbook could not be read") from exc
    if not raw_headers:
        raise HTTPException(status_code=400, detail="The workbook is empty")

    headers = [_normalize_import_header(value) for value in raw_headers]
    if len(set(filter(None, headers))) != len(list(filter(None, headers))):
        raise HTTPException(status_code=400, detail="The workbook contains duplicate column names")
    missing = [field for field in ("part_number", "name") if field not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    field_indexes = {field: headers.index(field) for field in PART_IMPORT_FIELDS if field in headers}
    custom_indexes = {header: index for index, header in enumerate(headers) if header.startswith("custom_")}
    normalized_rows: list[dict] = []
    errors: list[dict] = []
    seen_numbers: set[str] = set()
    seen_barcodes: set[str] = set()
    total_rows = 0
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(value not in (None, "") for value in row):
            continue
        total_rows += 1
        values = {field: row[index] if index < len(row) else None for field, index in field_indexes.items()}
        custom_fields = {
            field.removeprefix("custom_"): row[index]
            for field, index in custom_indexes.items()
            if index < len(row) and row[index] not in (None, "")
        }
        part_number = str(values.get("part_number") or "").strip()
        name = str(values.get("name") or "").strip()
        row_errors: list[str] = []
        if not part_number:
            row_errors.append("part_number is required")
        if not name:
            row_errors.append("name is required")
        if part_number and part_number in seen_numbers:
            row_errors.append("duplicate part_number in file")
        if part_number:
            seen_numbers.add(part_number)
        barcode = str(values.get("barcode") or "").strip() or None
        if barcode and barcode in seen_barcodes:
            row_errors.append("duplicate barcode in file")
        if barcode:
            seen_barcodes.add(barcode)
            barcode_owner = db.scalar(select(Part).where(Part.barcode == barcode))
            if barcode_owner and barcode_owner.part_number != part_number:
                row_errors.append("barcode already belongs to another item")
        tracking_mode = str(values.get("tracking_mode") or "none").strip().lower()
        if tracking_mode not in {"none", "batch", "serial"}:
            row_errors.append("tracking_mode must be none, batch, or serial")
        try:
            default_cost = _parse_non_negative_number(values.get("default_cost"), float, "default_cost")
            safety_stock = _parse_non_negative_number(values.get("safety_stock"), int, "safety_stock")
            min_stock = _parse_non_negative_number(values.get("min_stock"), int, "min_stock")
        except (TypeError, ValueError) as exc:
            row_errors.append(str(exc))
            default_cost, safety_stock, min_stock = 0.0, 0, 0

        if row_errors:
            errors.append({"row": row_number, "part_number": part_number or None, "messages": row_errors})
            continue
        normalized_rows.append(
            {
                "row_number": row_number,
                "part_number": part_number,
                "name": name,
                "category": str(values.get("category") or "").strip() or None,
                "barcode": barcode,
                "item_type": str(values.get("item_type") or "stock").strip() or "stock",
                "tracking_mode": tracking_mode,
                "is_active": str(values.get("is_active") or "true").strip().lower() not in {"false", "0", "no"},
                "custom_fields": custom_fields,
                "english_name": str(values.get("english_name") or "").strip() or None,
                "machine_type": str(values.get("machine_type") or "").strip() or None,
                "unit": str(values.get("unit") or "pcs").strip() or "pcs",
                "default_cost": default_cost,
                "safety_stock": safety_stock,
                "min_stock": min_stock,
                "supplier": str(values.get("supplier") or "").strip() or None,
                "image_url": str(values.get("image_url") or "").strip() or None,
                "notes": str(values.get("notes") or "").strip() or None,
            }
        )

    part_numbers = [row["part_number"] for row in normalized_rows]
    existing_numbers = set(
        db.scalars(select(Part.part_number).where(Part.part_number.in_(part_numbers))).all()
    ) if part_numbers else set()
    batch = ImportBatch(
        import_type="parts",
        filename=Path(file.filename).name,
        file_sha256=file_hash,
        status="ready" if not errors else "invalid",
        total_rows=total_rows,
        valid_rows=len(normalized_rows),
        error_rows=len(errors),
        created_count=sum(1 for number in part_numbers if number not in existing_numbers),
        updated_count=sum(1 for number in part_numbers if number in existing_numbers),
        payload_json=json.dumps(normalized_rows, ensure_ascii=False),
        errors_json=json.dumps(errors, ensure_ascii=False),
        created_by=actor.user_id,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return _import_batch_read(batch)


@router.post("/imports/parts/{batch_id}/commit", response_model=ImportBatchRead)
def commit_parts_import(
    batch_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.import_type != "parts":
        raise HTTPException(status_code=404, detail="Import batch not found")
    if batch.status == "committed":
        return _import_batch_read(batch)
    if batch.status != "ready":
        raise HTTPException(status_code=409, detail="Import batch has validation errors")

    rows = json.loads(batch.payload_json or "[]")
    created = 0
    updated = 0
    for row in rows:
        data = {key: value for key, value in row.items() if key != "row_number"}
        item = db.scalar(select(Part).where(Part.part_number == data["part_number"]))
        if item:
            for key, value in data.items():
                setattr(item, key, value)
            updated += 1
        else:
            db.add(Part(**data))
            created += 1
    batch.status = "committed"
    batch.created_count = created
    batch.updated_count = updated
    batch.committed_at = datetime.utcnow()
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return _import_batch_read(batch)


@router.get("/imports/parts", response_model=list[ImportBatchRead])
def list_parts_imports(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    batches = db.scalars(
        select(ImportBatch).where(ImportBatch.import_type == "parts").order_by(ImportBatch.id.desc()).offset(skip).limit(limit)
    ).all()
    return [_import_batch_read(batch) for batch in batches]


@router.post("/imports/opening-inventory/preview", response_model=ImportBatchRead)
async def preview_opening_inventory_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    content = await file.read(settings.max_import_upload_bytes + 1)
    if len(content) > settings.max_import_upload_bytes:
        raise HTTPException(status_code=413, detail="Import file exceeds the configured upload limit")
    file_hash = sha256(content).hexdigest()
    existing_batch = db.scalar(
        select(ImportBatch)
        .where(ImportBatch.import_type == "opening_inventory", ImportBatch.file_sha256 == file_hash)
        .order_by(ImportBatch.id.desc())
    )
    if existing_batch:
        return _import_batch_read(existing_batch)

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        worksheet = workbook.active
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="The uploaded workbook could not be read") from exc
    if not raw_headers:
        raise HTTPException(status_code=400, detail="The workbook is empty")
    headers = [_normalize_import_header(value) for value in raw_headers]
    missing = [field for field in ("part_number", "warehouse", "quantity") if field not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    indexes = {field: headers.index(field) for field in ("part_number", "warehouse", "quantity", "unit_cost", "notes") if field in headers}
    parts = {part.part_number: part for part in db.scalars(select(Part)).all()}
    warehouses = {warehouse.name.strip().lower(): warehouse for warehouse in db.scalars(select(Warehouse)).all()}
    normalized_rows: list[dict] = []
    errors: list[dict] = []
    seen_pairs: set[tuple[str, str]] = set()
    total_rows = 0
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(value not in (None, "") for value in row):
            continue
        total_rows += 1
        values = {field: row[index] if index < len(row) else None for field, index in indexes.items()}
        part_number = str(values.get("part_number") or "").strip()
        warehouse_name = str(values.get("warehouse") or "").strip()
        part = parts.get(part_number)
        warehouse = warehouses.get(warehouse_name.lower())
        row_errors: list[str] = []
        if not part:
            row_errors.append("part_number does not exist in this organization")
        if not warehouse:
            row_errors.append("warehouse does not exist in this organization")
        elif warehouse_is_vehicle(db, warehouse):
            row_errors.append(
                "opening inventory cannot post to a vehicle; use replenishment and engineer receipt"
            )
        pair = (part_number, warehouse_name.lower())
        if part_number and warehouse_name and pair in seen_pairs:
            row_errors.append("duplicate part and warehouse in file")
        seen_pairs.add(pair)
        try:
            quantity = int(values.get("quantity"))
            if quantity <= 0:
                raise ValueError("quantity must be greater than zero")
        except (TypeError, ValueError) as exc:
            row_errors.append(str(exc) if str(exc) else "quantity must be a whole number")
            quantity = 0
        try:
            unit_cost = _parse_non_negative_number(values.get("unit_cost"), float, "unit_cost")
        except (TypeError, ValueError) as exc:
            row_errors.append(str(exc))
            unit_cost = 0.0
        if row_errors:
            errors.append({"row": row_number, "part_number": part_number or None, "messages": row_errors})
            continue
        current_quantity = get_stock_quantity(db, part.id, warehouse.id)
        normalized_rows.append(
            {
                "row_number": row_number,
                "part_id": part.id,
                "part_number": part.part_number,
                "part_name": part.name,
                "warehouse_id": warehouse.id,
                "warehouse": warehouse.name,
                "quantity": quantity,
                "current_quantity": current_quantity,
                "projected_quantity": current_quantity + quantity,
                "unit_cost": unit_cost,
                "notes": str(values.get("notes") or "").strip() or None,
            }
        )

    batch = ImportBatch(
        import_type="opening_inventory",
        filename=Path(file.filename).name,
        file_sha256=file_hash,
        status="ready" if not errors else "invalid",
        total_rows=total_rows,
        valid_rows=len(normalized_rows),
        error_rows=len(errors),
        created_count=len(normalized_rows),
        updated_count=0,
        payload_json=json.dumps(normalized_rows, ensure_ascii=False),
        errors_json=json.dumps(errors, ensure_ascii=False),
        created_by=actor.user_id,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return _import_batch_read(batch)


@router.post("/imports/opening-inventory/{batch_id}/commit", response_model=ImportBatchRead)
def commit_opening_inventory_import(
    batch_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    begin_inventory_write(db)
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.import_type != "opening_inventory":
        raise HTTPException(status_code=404, detail="Import batch not found")
    if batch.status == "committed":
        return _import_batch_read(batch)
    if batch.status != "ready":
        raise HTTPException(status_code=409, detail="Import batch has validation errors")
    rows = json.loads(batch.payload_json or "[]")
    part_ids = sorted({row["part_id"] for row in rows})
    if part_ids:
        locked_parts = db.scalars(
            select(Part).where(Part.id.in_(part_ids)).order_by(Part.id).with_for_update()
        ).all()
        if len(locked_parts) != len(part_ids):
            raise HTTPException(status_code=409, detail="An opening inventory part no longer exists")
    for row in rows:
        warehouse = db.get(Warehouse, row["warehouse_id"])
        if not warehouse or warehouse_is_vehicle(db, warehouse):
            raise HTTPException(
                status_code=409,
                detail="Opening inventory can only be committed to non-vehicle warehouses",
            )
        db.add(
            InventoryTransaction(
                part_id=row["part_id"],
                transaction_type=TransactionType.INBOUND,
                quantity=row["quantity"],
                to_warehouse_id=row["warehouse_id"],
                user_id=actor.user_id,
                unit_cost=row["unit_cost"],
                notes=f"Opening inventory import #{batch.id}. {row.get('notes') or ''}".strip(),
            )
        )
    batch.status = "committed"
    batch.created_count = len(rows)
    batch.updated_count = 0
    batch.committed_at = datetime.utcnow()
    db.add(batch)
    _audit(
        db,
        actor,
        "opening_inventory_committed",
        "import_batch",
        batch.id,
        {"rows": len(rows), "warehouse_ids": sorted({row["warehouse_id"] for row in rows})},
    )
    db.commit()
    db.refresh(batch)
    return _import_batch_read(batch)


@router.get("/imports/opening-inventory", response_model=list[ImportBatchRead])
def list_opening_inventory_imports(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=100),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    batches = db.scalars(
        select(ImportBatch)
        .where(ImportBatch.import_type == "opening_inventory")
        .order_by(ImportBatch.id.desc())
        .offset(skip)
        .limit(limit)
    ).all()
    return [_import_batch_read(batch) for batch in batches]


@router.post("/import/parts.xlsx")
async def import_parts_excel(
    file: UploadFile = File(...), db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read(settings.max_import_upload_bytes + 1)
    if len(content) > settings.max_import_upload_bytes:
        raise HTTPException(status_code=413, detail="Import file exceeds the configured upload limit")
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0

    for row_number, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        part_number = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else part_number
        unit = str(row[2]).strip() if row[2] else "pcs"
        default_cost = float(row[3] or 0)
        safety_stock = int(row[4] or 0)
        supplier = str(row[5]).strip() if row[5] else None
        notes = str(row[6]).strip() if row[6] else None

        item = db.scalar(select(Part).where(Part.part_number == part_number))
        if item:
            item.name = name
            item.unit = unit
            item.default_cost = default_cost
            item.safety_stock = safety_stock
            item.supplier = supplier
            item.notes = notes
            updated += 1
        else:
            db.add(
                Part(
                    part_number=part_number,
                    name=name,
                    unit=unit,
                    default_cost=default_cost,
                    safety_stock=safety_stock,
                    supplier=supplier,
                    notes=notes,
                )
            )
            created += 1
    db.commit()
    return {"created": created, "updated": updated}


@router.get("/export/work-orders.xlsx")
def export_work_orders_excel(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    rows = db.scalars(select(WorkOrder).order_by(WorkOrder.id.asc())).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "WorkOrders"
    ws.append(
        [
            "wo_number",
            "ticket_number",
            "schedule_date",
            "outlet_name",
            "job_type",
            "description",
            "address",
            "city",
            "state",
            "zip",
            "contact_phone",
            "status",
            "revenue",
            "labor_cost",
            "assigned_user_id",
            "engineer_id",
        ]
    )
    for row in rows:
        ws.append(
            [
                row.wo_number,
                row.ticket_number,
                row.schedule_date.isoformat() if row.schedule_date else None,
                row.outlet_name,
                row.job_type,
                row.description,
                row.address,
                row.city,
                row.state,
                row.zip,
                row.contact_phone,
                row.status,
                row.revenue,
                row.labor_cost,
                row.assigned_user_id,
                row.engineer_id,
            ]
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=work-orders.xlsx"},
    )


@router.post("/import/work-orders.xlsx")
async def import_work_orders_excel(
    file: UploadFile = File(...), db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    content = await file.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    updated = 0

    for row in rows:
        if not row:
            continue
        wo_number = str(row[0]).strip() if row[0] else None
        ticket_number = str(row[1]).strip() if row[1] else None
        if not wo_number and not ticket_number:
            continue
        if not ticket_number:
            ticket_number = wo_number
        if not wo_number:
            wo_number = ticket_number

        item = db.scalar(select(WorkOrder).where(WorkOrder.ticket_number == ticket_number))
        if not item and wo_number:
            item = db.scalar(select(WorkOrder).where(WorkOrder.wo_number == wo_number))

        schedule_date = None
        if row[2]:
            if isinstance(row[2], datetime):
                schedule_date = row[2].date()
            else:
                schedule_date = datetime.fromisoformat(str(row[2])).date()

        payload = {
            "wo_number": wo_number,
            "ticket_number": ticket_number,
            "schedule_date": schedule_date,
            "outlet_name": str(row[3]).strip() if row[3] else None,
            "store_name": str(row[3]).strip() if row[3] else None,
            "job_type": str(row[4]).strip() if row[4] else None,
            "description": str(row[5]).strip() if row[5] else None,
            "problem_description": str(row[5]).strip() if row[5] else None,
            "address": str(row[6]).strip() if row[6] else None,
            "city": str(row[7]).strip() if row[7] else None,
            "state": str(row[8]).strip() if row[8] else None,
            "zip": str(row[9]).strip() if row[9] else None,
            "contact_phone": str(row[10]).strip() if row[10] else None,
            "status": str(row[11]).strip() if row[11] else "open",
            "revenue": float(row[12] or 0),
            "labor_cost": float(row[13] or 0),
            "assigned_user_id": int(row[14]) if row[14] else None,
            "engineer_id": int(row[15]) if row[15] else None,
        }
        if payload["status"].strip().upper() in {"COMPLETED", "PENDING_APPROVAL", "APPROVAL_REJECTED"}:
            raise HTTPException(status_code=400, detail=f"Row {row_number}: terminal status requires the completion workflow")
        _require_tenant_user(db, payload["assigned_user_id"], "assigned_user_id")
        _require_tenant_user(db, payload["engineer_id"], "engineer_id")
        if payload["assigned_user_id"] and not payload["engineer_id"]:
            payload["engineer_id"] = payload["assigned_user_id"]
        if payload["engineer_id"] and not payload["assigned_user_id"]:
            payload["assigned_user_id"] = payload["engineer_id"]

        if item:
            for key, value in payload.items():
                setattr(item, key, value)
            updated += 1
        else:
            db.add(WorkOrder(**payload))
            created += 1

    db.commit()
    return {"created": created, "updated": updated}


@router.get("/dashboard/engineers/{user_id}", response_model=EngineerDashboard)
def engineer_dashboard(
    user_id: int,
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.ENGINEER)
    if actor.role == UserRole.ENGINEER and actor.user_id != user_id:
        raise HTTPException(status_code=403, detail="Technicians can only view their own dashboard")
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    open_work_orders = db.scalar(
        select(func.count(WorkOrder.id)).where(
            WorkOrder.assigned_user_id == user_id,
            WorkOrder.status != "completed",
        )
    )
    completed_work_orders = db.scalar(
        select(func.count(WorkOrder.id)).where(
            WorkOrder.assigned_user_id == user_id,
            WorkOrder.status == "completed",
        )
    )
    van_inventory = get_employee_van_inventory(db, user_id)
    low_stock_items = sum(1 for item in van_inventory if item.is_low_stock)

    return EngineerDashboard(
        user_id=user.id,
        user_name=user.name,
        open_work_orders=open_work_orders or 0,
        completed_work_orders=completed_work_orders or 0,
        van_low_stock_items=low_stock_items,
        van_inventory=van_inventory,
    )


@router.get("/dashboard/admin/warehouses", response_model=AdminWarehouseDashboard)
def admin_warehouse_dashboard(
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    warehouses = db.scalars(select(Warehouse).order_by(Warehouse.id.asc())).all()
    all_balances = get_stock_balances(db)
    total_parts = db.scalar(select(func.count(Part.id))) or 0

    warehouse_rows: list[WarehouseSummary] = []
    total_low_stock = 0
    for warehouse in warehouses:
        rows = [balance for balance in all_balances if balance.warehouse_id == warehouse.id]
        low_stock_items = sum(1 for row in rows if row.is_low_stock)
        total_low_stock += low_stock_items

        warehouse_rows.append(
            WarehouseSummary(
                warehouse_id=warehouse.id,
                warehouse_name=warehouse.name,
                assigned_user_id=warehouse.assigned_user_id,
                assigned_user_name=warehouse.assigned_user.name if warehouse.assigned_user else None,
                total_sku=sum(1 for row in rows if row.quantity > 0),
                total_quantity=sum(row.quantity for row in rows),
                low_stock_items=low_stock_items,
            )
        )

    return AdminWarehouseDashboard(
        total_warehouses=len(warehouse_rows),
        total_parts=total_parts,
        total_low_stock_items=total_low_stock,
        warehouses=warehouse_rows,
    )


@router.get("/inventory/low-stock-alerts", response_model=list[LowStockAlert])
def low_stock_alerts(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER, UserRole.WAREHOUSE)
    rows = get_stock_balances(db)
    alerts = [
        LowStockAlert(
            part_id=row.part_id,
            part_number=row.part_number,
            part_name=row.part_name,
            warehouse_id=row.warehouse_id,
            warehouse_name=row.warehouse_name,
            quantity=row.quantity,
            min_stock=max(row.safety_stock, db.get(Part, row.part_id).min_stock if db.get(Part, row.part_id) else 0),
        )
        for row in rows
        if row.is_low_stock
    ]
    return alerts[skip : skip + limit]


@router.get("/reports/abnormal-usage", response_model=list[AbnormalUsageRow])
def abnormal_usage_report(
    skip: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    db: Session = Depends(get_db),
    actor: Actor = Depends(get_current_actor),
):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    work_orders = db.scalars(select(WorkOrder).order_by(WorkOrder.id.asc())).all()
    if not work_orders:
        return []
    costs = sorted([get_work_order_parts_cost(db, wo.id) for wo in work_orders if get_work_order_parts_cost(db, wo.id) > 0])
    if costs:
        idx = min(len(costs) - 1, int(0.9 * (len(costs) - 1)))
        p90_cost = costs[idx]
    else:
        p90_cost = 0.0

    tech_avgs: dict[int, float] = {}
    tech_map: dict[int, list[float]] = {}
    for wo in work_orders:
        if not wo.engineer_id:
            continue
        tech_map.setdefault(wo.engineer_id, []).append(get_work_order_parts_cost(db, wo.id))
    for tech_id, arr in tech_map.items():
        tech_avgs[tech_id] = (sum(arr) / len(arr)) if arr else 0.0

    repeated_part_flags: dict[int, set[int]] = {}
    usage_rows = db.scalars(select(WorkOrderPart).order_by(WorkOrderPart.created_at.asc())).all()
    for row in usage_rows:
        if not row.user_id:
            continue
        repeated_part_flags.setdefault(row.user_id, set())
        # repeated same part usage in short period (7 days)
        recent_count = sum(
            1
            for r in usage_rows
            if r.user_id == row.user_id
            and r.part_id == row.part_id
            and r.created_at >= (row.created_at - timedelta(days=7))
            and r.created_at <= row.created_at
        )
        if recent_count >= 3:
            repeated_part_flags[row.user_id].add(row.work_order_id)

    results: list[AbnormalUsageRow] = []
    for wo in work_orders:
        parts_cost = get_work_order_parts_cost(db, wo.id)
        reasons: list[str] = []
        if parts_cost >= p90_cost and parts_cost > 0:
            reasons.append("Parts cost above 90th percentile")
        if wo.engineer_id and tech_avgs.get(wo.engineer_id, 0) > 0 and parts_cost > tech_avgs[wo.engineer_id] * 1.8:
            reasons.append("Technician parts usage above historical average")
        if wo.engineer_id and wo.id in repeated_part_flags.get(wo.engineer_id, set()):
            reasons.append("Repeated same-part usage in short period")
        if reasons:
            results.append(
                AbnormalUsageRow(
                    work_order_id=wo.id,
                    ticket_number=wo.ticket_number,
                    engineer_id=wo.engineer_id,
                    parts_cost=parts_cost,
                    revenue=wo.revenue,
                    severity="high" if len(reasons) > 1 else "medium",
                    reason="; ".join(reasons),
                )
            )
    return results[skip : skip + limit]


@router.get("/pilot/checklist")
def pilot_checklist(db: Session = Depends(get_db), actor: Actor = Depends(get_current_actor)):
    require_roles(actor, UserRole.ADMIN, UserRole.MANAGER)
    low_stock = low_stock_alerts(db=db, actor=actor)
    abnormal = abnormal_usage_report(db=db, actor=actor)
    return {
        "system_health": "ok",
        "total_users": db.scalar(select(func.count(User.id))) or 0,
        "total_work_orders": db.scalar(select(func.count(WorkOrder.id))) or 0,
        "total_parts": db.scalar(select(func.count(Part.id))) or 0,
        "total_inventory_transactions": db.scalar(select(func.count(InventoryTransaction.id))) or 0,
        "low_stock_alert_count": len(low_stock),
        "abnormal_usage_alert_count": len(abnormal),
    }
